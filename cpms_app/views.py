from django.shortcuts import render, get_object_or_404, redirect
import logging
from django.contrib.auth import get_user_model
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.forms import PasswordResetForm
from django.contrib import messages
from django.utils.dateparse import parse_date
from .models import Milestone, Project, Team,  ChatRoom, Message, TaskUpdate, Task
from django.contrib.auth.decorators import login_required
User = get_user_model()
from .models import Profile
from .models import Milestone
from django.http import JsonResponse
import plotly.express as px
import pandas as pd
from django.shortcuts import render
from django.http import HttpResponseRedirect
from django.http import HttpResponseForbidden
import re
from django.core.paginator import Paginator
from django.http import Http404
from datetime import date
from django.db.models import Count, F, ExpressionWrapper, DateField
import os
from django.db.models import Count, Prefetch
from django.utils.timezone import now
from django.utils import timezone
import json
import shutil
from django.contrib.auth.models import Permission
from django.http import JsonResponse
from django.db.models import Count
from .models import ActivityLog
from django.conf import settings
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponseForbidden
from .models import ChatRoom, Message, User ,Notification, TaskStep, TaskIssue
from .models import Notification
from .decorators import role_required,project_leader_required,team_member_required
from django.core.files.storage import FileSystemStorage
from itertools import chain
import openpyxl
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go
from django.db.models import Count, Avg
from django.contrib.auth.views import PasswordResetCompleteView
from django.urls import reverse_lazy
from django.db.models import Count, Q
from datetime import datetime, date
from rest_framework.views import APIView
from rest_framework.response import Response
from django.urls import reverse


def landing_page(request):
    if request.user.is_staff:  # Check if the user is an admin (staff)
        return admin_landing_page(request)
    else:
        return other_landing_page(request)  # Render for non-admin users
    
def other_landing_page(request):
    user = request.user
    # Project Section
    projects = Project.objects.filter(Q(assigned_team__members=user))
    total_projects=projects.count()
    completed_projects = projects.filter(Q(status='completed'))
    pending_projects = projects.filter(Q(status='pending'))
    ongoing_projects = projects.filter(status = "in_progress").order_by("-updated_at")

    user_teams = Team.objects.filter(members=request.user).count()

    recent_milestones = Milestone.objects.filter(Q(assigned_to=user) ,  Q(status__in=["pending", "in_progress"] )).select_related("project").distinct().order_by("created_at")[:5]
    # Q(updates__updated_by=user),
    # Task section
    milestones=Milestone.objects.filter(Q(assigned_to=user))
    total_milestones=milestones.count()
    ongoing_milestones = milestones.filter(assigned_to=user, status="in_progress")
    pending_milestones = milestones.filter(assigned_to=user, status="pending")
    completed_milestones = milestones.filter(assigned_to=user, status="completed")

    # Task section
    teams=Team.objects.filter(Q(members=user))


    # Context to pass to the template
    context = {
        "teams":teams,
        "ongoing_projects": ongoing_projects,  
        "projects":projects, 
        "recent_milestones": recent_milestones,
        "ongoing_milestones": ongoing_milestones,
        "pending_milestones": pending_milestones,
        "total_milestones":total_milestones,
        "milestones":milestones,
        "completed_milestones": completed_milestones,
        "user_teams":user_teams,
        "total_projects":total_projects,
        "completed_projects":completed_projects,
        "pending_projects":pending_projects,
    }
    return render(request, 'landing_page.html',context)
def admin_activity_view(request):
    # Get filter parameters
    search_query = request.GET.get('search', '')
    user_filter = request.GET.get('user', '')
    project_filter = request.GET.get('project', '')
    milestone_filter = request.GET.get('milestone', '')
    activity_type_filter = request.GET.get('activity_type', '')

    # Base query
    activities = ActivityLog.objects.all()

    # Apply search query filter
    if search_query:
        activities = activities.filter(
            Q(activity_description__icontains=search_query) |
            Q(activity_type__icontains=search_query)
        )

    # Apply user filter
    if user_filter:
        activities = activities.filter(user_id=user_filter)

    # Apply project filter
    if project_filter:
        activities = activities.filter(object_id=project_filter, activity_type__in=['create_project', 'edit_project', 'delete_project'])

    # Apply milestone filter
    if milestone_filter:
        activities = activities.filter(object_id=milestone_filter, activity_type__in=['milestone_created', 'milestone_updated', 'milestone_deleted'])

    # Apply activity type filter
    if activity_type_filter:
        activities = activities.filter(activity_type=activity_type_filter)

    # Order activities by timestamp
    activities = activities.order_by('-timestamp')

    # Get dropdown choices
    users = User.objects.all()
    projects = Project.objects.all()
    milestones = Milestone.objects.all()
    activity_choices = dict(ActivityLog.ACTIVITY_CHOICES)

    context = {
        'activities': activities,  # Send all activities (no pagination as DataTable handles it)
        'users': users,
        'projects': projects,
        'milestones': milestones,
        'activity_choices': activity_choices,
        'search_query': search_query,
        'user_filter': user_filter,
        'project_filter': project_filter,
        'milestone_filter': milestone_filter,
        'activity_type_filter': activity_type_filter,
    }

    return render(request, 'admin/activity_view.html', context)

def admin_landing_page(request):
    # User role check
    user = request.user
 
    # Statistics for Projects
    total_projects = Project.objects.count()
    ongoing_projects = Project.objects.filter(status='in_progress').count()
    completed_projects = Project.objects.filter(status='completed').count()
    pending_projects = Project.objects.filter(status='pending').count()

    # Statistics for Teams
    total_teams = Team.objects.count()
    teams_with_no_projects = Team.objects.filter(projects__isnull=True).count()
    teams_with_more_than_2_projects = (
    Team.objects.annotate(project_count=Count('projects'))
    .filter(project_count__gt=1)
    .count()
    )

    other_teams = total_teams - teams_with_no_projects - teams_with_more_than_2_projects

    # Statistics for Users
    total_users = User.objects.count()
    users_without_teams = User.objects.filter(teams__isnull=True).count()
    users_without_milestones = User.objects.filter(milestones__isnull=True).count()
    users_in_more_than_2_teams = User.objects.annotate(team_count=Count('teams')).filter(team_count__gt=1).count()
    admin_users= User.objects.filter(role=['admin','super_admin']).count()

  
    recent_activities = ActivityLog.objects.filter(
    user=request.user,
   
).order_by('-timestamp')[:5]  # Limit to the 5 most recent activities

    # Context preparation
    context = {
        'total_projects': total_projects,
        'ongoing_projects': ongoing_projects,
        'completed_projects': completed_projects,
        'pending_projects': pending_projects,
        'total_teams': total_teams,
        'teams_with_no_projects': teams_with_no_projects,
        'teams_with_more_than_2_projects': teams_with_more_than_2_projects,
        'other_teams': other_teams,
        'total_users': total_users,
        'users_without_teams': users_without_teams,
        'users_without_milestones': users_without_milestones,
        'users_in_more_than_2_teams': users_in_more_than_2_teams,
        'recent_activities':recent_activities,
        
    }

    return render(request, 'admin/landing_page.html', context)

@role_required(['super_admin','admin','staff'])
def report_dashboard(request):
    projects = Project.objects.all()
    teams = Team.objects.all()

    # Filtering
    project_id = request.GET.get("project")
    team_id = request.GET.get("team")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    filtered_projects = projects
    if project_id:
        filtered_projects = filtered_projects.filter(id=project_id)
    if team_id:
        filtered_projects = filtered_projects.filter(assigned_team__id=team_id)
    if start_date:
        filtered_projects = filtered_projects.filter(start_date__gte=start_date)
    if end_date:
        filtered_projects = filtered_projects.filter(end_date__lte=end_date)
    for project in projects:
            project.length = project.project_length()
            project.remaining_days = project.days_left()
            # Calculate progress
            project.progress = project.calculate_progress()
    context = {
        "projects": projects,
        "teams": teams,
        "filtered_projects": filtered_projects,
    }
    return render(request, "admin/report_dashboard.html", context)

@role_required(['super_admin','admin','staff'])
def generate_filtered_excel(request):
    projects = Project.objects.all()

    # Apply filters as in `report_dashboard`
    project_id = request.GET.get("project")
    team_id = request.GET.get("team")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if project_id:
        projects = projects.filter(id=project_id)
    if team_id:
        projects = projects.filter(assigned_team__id=team_id)
    if start_date:
        projects = projects.filter(start_date__gte=start_date)
    if end_date:
        projects = projects.filter(end_date__lte=end_date)

    # Create Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Filtered Reports"

    # Header
    headers = ["Project Name", "Team Name", "Start Date", "End Date", "Progress"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Data
    for row_num, project in enumerate(projects, 2):
        sheet[f"A{row_num}"] = project.name
        sheet[f"B{row_num}"] = project.assigned_team.name if project.assigned_team else "No Team"
        sheet[f"C{row_num}"] = project.start_date
        sheet[f"D{row_num}"] = project.end_date
        sheet[f"E{row_num}"] = project.progress

    # Response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=filtered_report.xlsx"
    workbook.save(response)
    return response


    # Create a new workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Project Overview"

    # Define the report header
    sheet['A1'] = "Project Name"
    sheet['B1'] = "Start Date"
    sheet['C1'] = "End Date"
    sheet['D1'] = "Status"
    sheet['E1'] = "Total Tasks"
    sheet['F1'] = "Completed Tasks"
    sheet['G1'] = "Pending Tasks"
    sheet['H1'] = "Progress (%)"
    sheet['I1'] = "Assigned Team"  # New header for assigned team

    # Fetch all projects
    projects = Project.objects.all()
    row = 2  # Start from the second row for data

    for project in projects:
        milestones = project.milestones.all()

        # Calculate statistics related to milestones
        total_milestones = milestones.count()
        completed_milestones = milestones.filter(status='completed').count()
        pending_milestones = milestones.filter(status='pending').count()
        progress = project.calculate_progress()
        assigned_team = project.assigned_team.name if project.assigned_team else "N/A"  # Fetch the team name

        # Write data to excel
        sheet[f'A{row}'] = project.name
        sheet[f'B{row}'] = project.start_date
        sheet[f'C{row}'] = project.end_date
        sheet[f'D{row}'] = project.status
        sheet[f'E{row}'] = total_milestones
        sheet[f'F{row}'] = completed_milestones
        sheet[f'G{row}'] = pending_milestones
        sheet[f'H{row}'] = progress
        sheet[f'I{row}'] = assigned_team  # Add the assigned team name

        row += 1  # Move to the next row

    # Generate response for Excel download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=project_overview.xlsx'
    workbook.save(response)  # Write the Excel file to the response
    return response

@login_required
def dashboard(request):
    projects = Project.objects.all()
    teams = Team.objects.all()

    # Determine the active tab based on the GET parameters
    active_tab = "project_status"  # Default tab
    if request.GET.get('project_id'):
        active_tab = "milestones_by_status"
    elif request.GET.get('team_id'):
        active_tab = "team_contribution"

    # Get the selected project and generate the chart
    selected_project_id = request.GET.get('project_id')
    project_milestones_chart = milestones_status_chart(project_id=selected_project_id) if selected_project_id else milestones_status_chart()

    # Get the selected team and generate the chart
    selected_team_id = request.GET.get('team_id')
    team_contrib_chart = team_contribution_chart(team_id=selected_team_id) if selected_team_id else team_contribution_chart()
    
    # Context for the template
    context = {
        'teams': teams,
        'team_contrib_chart': team_contrib_chart,
        'projects': projects,
        'project_milestones_chart': project_milestones_chart,
        "project_status_chart": project_status_chart(),
        "milestones_status_chart": milestones_status_chart(),
        "team_contribution_chart": team_contribution_chart(),
        "project_progress_chart": project_progress_chart(),
        "remaining_days_chart": remaining_days_chart(),
        "delayed_projects_chart": delayed_projects_chart(),
        "effort_accuracy_chart": effort_accuracy_chart(),
        "milestones_completed_chart": milestones_completed_chart(),
        "team_milestone_allocation_chart": team_milestone_allocation_chart(),
        "active_tab": active_tab,  # Pass the active tab to the template
    }
    return render(request, "dashboard.html", context)



def project_status_chart():
    status_counts = Project.objects.values("status").annotate(count=Count("status"))
    labels = [entry["status"].replace("_", " ").capitalize() for entry in status_counts]
    values = [entry["count"] for entry in status_counts]

    fig = go.Figure(
        go.Pie(
            labels=labels,
            values=values,
            hole=0.3,
            textinfo="label+value",  # Display the label and the actual count
            textposition="inside"
        )
    )
    fig.update_layout(title="Project Status Distribution")
    return fig.to_html()

def milestones_status_chart(project_id=None):
    # Filter milestones by project_id if it's provided, else all milestones
    if project_id:
        milestone_counts = Milestone.objects.filter(project_id=project_id).values("status").annotate(count=Count("status"))
    else:
        milestone_counts = Milestone.objects.values("status").annotate(count=Count("status"))

    # Prepare data for the chart
    statuses = [entry["status"].capitalize() for entry in milestone_counts]
    counts = [entry["count"] for entry in milestone_counts]

    fig = go.Figure(
        go.Bar(
            x=statuses,
            y=counts,
            marker_color=["blue", "orange", "green"],  # You can add more or modify colors as needed
        )
    )
    fig.update_layout(
        title="Task Status Distribution", 
        xaxis_title="Status", 
        yaxis_title="Count"
    )
    
    return fig.to_html()

def team_contribution_chart(team_id=None):
    if team_id:
        team = Team.objects.get(id=team_id)
        members = team.members.all()
        member_milestone_counts = []
        total_milestones = Milestone.objects.filter(project__assigned_team=team).count()

        for member in members:
            milestone_count = Milestone.objects.filter(assigned_to=member).count()
            percentage = (milestone_count / total_milestones) * 100 if total_milestones else 0
            member_milestone_counts.append({
                'username': member.username,
                'milestone_count': milestone_count,
                'percentage': round(percentage, 2)
            })

        # Extract data for the chart
        usernames = [entry['username'] for entry in member_milestone_counts]
        milestone_counts = [entry['milestone_count'] for entry in member_milestone_counts]
        percentages = [entry['percentage'] for entry in member_milestone_counts]

        # Generate Bar Chart with Task Count and Percentage
        fig = go.Figure()

        # Add bars for milestone count
        fig.add_trace(go.Bar(
            x=usernames,
            y=milestone_counts,
            name='Task Count',
            text=milestone_counts,  # Show milestone count on the bars
            textposition='auto',
            hoverinfo='x+y',  # Show milestone count on hover
            marker=dict(color='rgb(55, 83, 109)')
        ))


        fig.update_layout(
            barmode='group',  # Group the bars side by side
            title="Team Members Task Contribution",
            xaxis_title="Team Member",
            yaxis_title="Task Count ",
            yaxis=dict(
                title="Task Count",
                range=[0, max(milestone_counts )]  # Adjust y-axis range
            )
        )

    else:
        fig = go.Figure()

    return fig.to_html()


 
def project_progress_chart():
    projects = Project.objects.all()
    names = [project.name for project in projects]
    progress_values = [project.calculate_progress()for project in projects]

    fig = go.Figure(go.Bar(x=names, y=progress_values, text=progress_values, textposition="auto"))
    fig.update_layout(
        title="Project Progress Status",
        xaxis_title="Project",
        yaxis_title="Progress (%)",
        yaxis=dict(range=[0, 100]),  # Limit to 0-100%
    )
    return fig.to_html()
def remaining_days_chart():
    projects = Project.objects.filter(end_date__gte=date.today())
    names = [project.name for project in projects]
    remaining_days = [project.days_left() for project in projects]

    fig = go.Figure(go.Bar(x=names, y=remaining_days, text=remaining_days, textposition="auto"))
    fig.update_layout(
        title="Remaining Days for Projects",
        xaxis_title="Project",
        yaxis_title="Days Left",
    )
    return fig.to_html()
def delayed_projects_chart():
    delayed_projects = Project.objects.filter(end_date__lt=date.today(), status__in=["pending", "in_progress"])
    names = [project.name for project in delayed_projects]
    delays = [abs(project.days_left()) for project in delayed_projects]

    fig = go.Figure(go.Bar(x=names, y=delays, text=delays, textposition="auto", marker_color="red"))
    fig.update_layout(
        title="Delayed Projects",
        xaxis_title="Project",
        yaxis_title="Days Delayed",
    )
    return fig.to_html()
def effort_accuracy_chart():
    milestones = Milestone.objects.filter(effort_estimation__isnull=False, actual_effort__isnull=False)
    milestone_names = [milestone.name for milestone in milestones]
    estimated_effort = [milestone.effort_estimation for milestone in milestones]
    actual_effort = [milestone.actual_effort for milestone in milestones]

    fig = go.Figure()
    fig.add_trace(go.Bar(x=milestone_names, y=estimated_effort, name="Estimated Effort", marker_color="blue"))
    fig.add_trace(go.Bar(x=milestone_names, y=actual_effort, name="Actual Effort", marker_color="orange"))
    fig.update_layout(
        title="Effort Accuracy",
        xaxis_title="Tasks",
        yaxis_title="Hours",
        barmode="group",
    )
    return fig.to_html()
def milestones_completed_chart():
    teams = Team.objects.all()
    team_names = []
    completion_rates = []

    for team in teams:
        # For each team, find all the milestones from the projects assigned to that team
        projects = Project.objects.filter(assigned_team=team)
        total_milestones = 0
        completed_milestones = 0

        # Iterate over projects assigned to the team and count completed milestones
        for project in projects:
            milestones = project.milestones.all()  # Fetch milestones related to the project
            total_milestones += milestones.count()  # Count the total number of milestones for the project
            completed_milestones += milestones.filter(status="completed").count()  # Count completed milestones

        if total_milestones > 0:
            team_names.append(team.name)
            completion_rates.append(round((completed_milestones / total_milestones) * 100, 2))

    fig = go.Figure(go.Bar(x=team_names, y=completion_rates, text=completion_rates, textposition="auto"))
    fig.update_layout(
        title="Task Completion Percentage by Team",
        xaxis_title="Team",
        yaxis_title="Completion Rate (%)",
        yaxis=dict(range=[0, 100]),
    )
    return fig.to_html()

def team_milestone_allocation_chart():
    teams = Team.objects.all()
    team_names = [team.name for team in teams]
    milestone_counts = []

    for team in teams:
        # For each team, find all milestones in projects assigned to that team
        projects = Project.objects.filter(assigned_team=team)
        total_milestones = 0
        for project in projects:
            milestones = project.milestones.all()  # Get milestones related to the project
            total_milestones += milestones.count()  # Count milestones assigned to team in the project

        milestone_counts.append(total_milestones)

    fig = go.Figure(go.Pie(labels=team_names, values=milestone_counts, hole=0.3))
    fig.update_layout(title="Task Allocation by Team")
    return fig.to_html()

@login_required
def user_dashboard(request):
    user = request.user

    # Get the user's teams
    user_teams = user.teams.all()

    # Get all projects associated with the user via team assignment
    relevant_projects = Project.objects.filter(assigned_team__in=user_teams).distinct()

    # Fetch the milestones
    assigned_milestones = Milestone.objects.filter(assigned_to=user)
    team_milestones = Milestone.objects.filter(assigned_group__in=user_teams).distinct()

    # Combine milestones
    relevant_milestones = list(chain(assigned_milestones, team_milestones))

    # Create a dictionary to group detailed milestone data by project
    project_milestones = {}
    today = now().date()
    for milestone in relevant_milestones:
        if milestone.project not in project_milestones:
            project_milestones[milestone.project] = []

        # Prepare milestone details
        status_display = dict(Milestone.STATUS_CHOICES).get(milestone.status, milestone.status)
        days_left = (milestone.end_date - today).days if milestone.end_date and milestone.end_date > today else None
        delayed_days = (today - milestone.end_date).days if milestone.end_date and milestone.end_date < today else None

        project_milestones[milestone.project].append({
            "id": milestone.id,
            "name": milestone.name,
            "description": milestone.description or "No description provided",
            "start_date": milestone.start_date,
            "end_date": milestone.end_date,
            'status': status_display,
            "days_left": f"{days_left} days " if days_left else "No days",
            "delayed_days": f"{delayed_days} days delayed" if delayed_days else None,
            "assigned_to": [user.username for user in milestone.assigned_to.all()],  # List of usernames
        })

    context = {
        'user': user,
        'projects': relevant_projects,
        'project_milestones': project_milestones,
    }
    return render(request, 'users/users_dashboard.html', context)
@role_required(['super_admin','admin','staff'])
def notifications_view(request):
    notifications = request.user.notifications.filter(is_read=False).order_by('-timestamp')
    return render(request, 'notifications.html', {'notifications': notifications})


@role_required(['super_admin','admin','staff'])
def mark_notification_read(request, notification_id):
    if request.method == 'POST':
        try:
            notification = Notification.objects.get(id=notification_id, recipient=request.user)
            notification.is_read = True
            notification.save()
            return redirect(f"{request.META.get('HTTP_REFERER', '/')}")
        except Notification.DoesNotExist:
            return redirect(f"{request.META.get('HTTP_REFERER', '/')}")
    return redirect(f"{request.META.get('HTTP_REFERER', '/')}")


@role_required(['super_admin', 'admin', 'staff'])
def send_message(request, room_id):
    # Get the chat room
    chat_room = get_object_or_404(ChatRoom, id=room_id)
    # Check if the user is a member of the project
    project = chat_room.project.assigned_team
    
    if request.user not in project.members.all():
        messages.error(request, 'You are not a member of this project.')
        return redirect('chat_room_detail', room_id=room_id)

    if request.method == 'POST':
        content = request.POST.get('content')
        reply_to_id = request.POST.get('reply_to')
        file = request.FILES.get('file')

        # Check if the content is not empty or there is no file uploaded
        if not content.strip() and not file:
            messages.error(request, 'Please write a message or upload a file.')
            return redirect('chat_room_detail', room_id=room_id)

        # Handle Mentions (extract usernames from @username syntax)
        mentioned_usernames = re.findall(r'@(\w+)', content)
        mentions = User.objects.filter(username__in=mentioned_usernames)

        # Create a new message object
        message = Message(
            user=request.user,
            room=chat_room,
            content=content,
            timestamp=now(),
        )

        # If the message is a reply, link the reply
        if reply_to_id:
            try:
                reply_to_message = Message.objects.get(id=reply_to_id)
                message.reply_to = reply_to_message
            except Message.DoesNotExist:
                messages.error(request, 'The message you are replying to does not exist.')
                return redirect('chat_room_detail', room_id=room_id)

        # Handle file upload (if present)
        if file:
            try:
                message.file = file
            except ValidationError as e:
                messages.error(request, f'File upload error: {str(e)}')
                return redirect('chat_room_detail', room_id=room_id)

        # Save the message
        message.save()

        # Log this activity
        ActivityLog.objects.create(
            user=request.user,  # The user who sent the message
            activity_type='message_sent',  # Define the activity type
            activity_description=f'Message sent in chat room "{chat_room.name}".',  # Description of the action
            object_name=chat_room.name,  # Chat room name where the message was sent
            object_id=chat_room.id,  # ID of the chat room
            timestamp=now(),  # Timestamp of the action
        )

        # Notify mentioned users
        for user in mentions:
            Notification.objects.create(
                recipient=user,
                message=f"You were mentioned in a chat by {request.user.username}."
            )

        # Redirect back to the chat room with a success message
        messages.success(request, 'Message sent successfully.')
        return redirect('chat_room_detail', room_id=room_id)

    # If it's not a POST request, redirect to the chat room
    return redirect('chat_room_detail', room_id=room_id)


@role_required(['super_admin','admin','staff'])
def edit_message(request, room_id, message_id):
    # Get the room and message to edit
    chat_room = get_object_or_404(ChatRoom, id=room_id)
    message = get_object_or_404(Message, id=message_id)

    # Check if the user is allowed to edit the message (creator or admin)
    if message.user == request.user or request.user.is_staff:
        if request.method == 'POST':
            # Get new content and file from the form
            new_content = request.POST.get('content')
            new_file = request.FILES.get('file')

            # Update message content
            old_content = message.content  # Save original content to compare later
            message.content = new_content
            if new_file:
                message.file = new_file  # Update the file if a new one is provided

            message.save()

            # Log this activity for editing the message
            ActivityLog.objects.create(
                user=request.user,
                activity_type='message_edited',  # Define activity type
                activity_description=f'Message edited in chat room "{chat_room.name}". Old content: "{old_content}".',
                object_name=chat_room.name,
                object_id=chat_room.id,
                timestamp=now(),
            )

            messages.success(request, 'Message updated successfully!')
            return redirect('chat_room_detail', room_id=room_id)
        else:
            messages.error(request, 'Invalid request method.')
    else:
        messages.error(request, 'You are not authorized to edit this message.')

    return redirect('chat_room_detail', room_id=room_id)


@role_required(['super_admin','admin','staff'])
def delete_message(request, message_id):
    # Get the message object using the message_id
    message = get_object_or_404(Message, id=message_id)

    # Check if the user has permission to delete the message (optional)
    if message.user != request.user and not request.user.is_staff:
        return HttpResponse("You do not have permission to delete this message.", status=403)

    # Log the activity for deleting the message
    ActivityLog.objects.create(
        user=request.user,
        activity_type='message_deleted',  # Define activity type
        activity_description=f'Message deleted in chat room "{message.room.name}".',
        object_name=message.room.name,  # The chat room where the message was deleted
        object_id=message.room.id,  # ID of the chat room
        timestamp=now(),
    )

    # Delete the message
    room_id = message.room.id
    message.delete()

    # Redirect to the chat room or message list after deletion
    return redirect('chat_room_detail', room_id=room_id)



def chat_room_detail(request, room_id):
    chat_room = get_object_or_404(ChatRoom, project=room_id)
    
    messages_list = Message.objects.filter(room=chat_room).order_by('timestamp')
    project = get_object_or_404(Project, id=room_id)
 
    team = chat_room.project.assigned_team
    team_members = team.members.select_related('profile') if team else []
    
    # Log activity when entering the chat room (viewing)
    ActivityLog.objects.create(
        user=request.user,
        activity_type='view_chat_room',
        activity_description=f'Viewed chat room "{chat_room.name}"',
        object_name=chat_room.name,
        object_id=chat_room.id,
        timestamp=now()
    )

    # Handling message edits
    if request.method == 'POST' and 'edit_message_id' in request.POST:
        message_id = request.POST.get('edit_message_id')
        new_content = request.POST.get('new_content')
        new_file = request.FILES.get('new_file')

        # Get the message to be edited
        message = get_object_or_404(Message, id=message_id)

        # Ensure the user editing the message is the message's creator or an admin
        if message.user == request.user or request.user.is_staff:
            # Log the activity before the update to track the content
            old_content = message.content  # Save old content before it is changed
            
            # Update the content
            message.content = new_content
            if new_file:
                message.file = new_file  # Update the file if provided
            message.save()

            # Log the edit activity
            ActivityLog.objects.create(
                user=request.user,
                activity_type='edit_message',
                activity_description=f'Edited message in chat room "{chat_room.name}". Old content: "{old_content}".',
                object_name=chat_room.name,
                object_id=chat_room.id,
                timestamp=now()
            )

            # Show success message and redirect
            messages.success(request, 'Message updated successfully!')
            return redirect('chat_room_detail', room_id=room_id)

        # If the user doesn't have permission to edit, show an error
        messages.error(request, 'You do not have permission to edit this message.')
        return redirect('chat_room_detail', room_id=room_id)

    # Return the rendered template with necessary context
    return render(request, 'chats/chat_room_detail.html', {
        'chat_room': chat_room,
        'messagess': messages_list,
        'team_members': team_members,
        'user': request.user,
        'project': project,
    })


def chatroom(request,room_id):
    return render(request,'chats/chatroom.html')


def gantt_chart(request, project_id):
    # Fetch the project and its milestones
    project = get_object_or_404(Project, id=project_id)
    milestones = project.milestones.all()  # Get all milestones related to the project

    # Check if there are no milestones
    if not milestones:
        return render(request, 'index.html', {
            'chart_html': None,
            'project': project,
            'message': "No milestones available for this project."
        })
    
    # Prepare data for the Gantt chart
    data = []
    for milestone in milestones:
        # Handle many-to-many assignment
        assigned_users = ', '.join([user.username for user in milestone.assigned_to.all()]) if milestone.assigned_to.exists() else 'Unassigned'
        
        data.append({
            'Task': milestone.name,
            'Start': milestone.start_date,
            'Finish': milestone.end_date,
            'Assigned To': assigned_users,
        })
    
    # Convert data to a pandas DataFrame
    df = pd.DataFrame(data)
    
    # Create the Gantt chart
    fig = px.timeline(df, x_start="Start", x_end="Finish", y="Task", color="Assigned To", title=f"{project.name}")
    fig.update_yaxes(categoryorder="total ascending")  # Sort milestones by start date
    
    # Convert the chart to HTML
    chart_html = fig.to_html(full_html=False)
    
    # Render the template with the Gantt chart
    return render(request, 'index.html', {'chart_html': chart_html, 'project': project})


def custom_login(request):
    # Check if 'next' is present in the URL parameters
    next_url = request.GET.get('next', None)

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        try:
            # Try to find the user manually first
            user = User.objects.get(username=username)

            # Check if user is inactive before authenticating
            if not user.is_active:
                # Log the failed login attempt due to inactive user status
                ActivityLog.objects.create(
                    user=user,
                    activity_type='failed_login',
                    activity_description=f'Login attempt blocked due to inactive status for username {username}',
                    object_name=username,
                    object_id=user.id,
                    timestamp=now()
                )
                messages.error(request, 'Your account is inactive. Please contact an administrator.')
                return redirect('login')  # Prevent login for inactive users

            # Proceed with authentication if user is active
          
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                # Log successful login activity
                ActivityLog.objects.create(
                    user=user,
                    activity_type='login',
                    activity_description='Successful login',
                    object_name=user.username,
                    object_id=user.id,
                    timestamp=now()
                )

                # Redirect based on 'next' or default to home page
                return redirect(next_url or 'home')

            else:
                # Handle failed login due to invalid credentials
                if user is not None:
                    ActivityLog.objects.create(
                        user=None,  # No user associated with invalid credentials
                        activity_type='failed_login',
                        activity_description=f'Failed login attempt for username {username} - Invalid credentials',
                        object_name=username,
                        object_id=None,
                        timestamp=now()
                    )
                messages.error(request, 'Invalid username or password')
                return redirect('login')

        except User.DoesNotExist:
            messages.error(request, 'Invalid username or password')
            return redirect('login')

    return render(request, 'users/login.html')
  
@login_required
def profile(request):
    
    # Get or create the profile for the logged-in user
    profile, created = Profile.objects.get_or_create(user=request.user)

    # Log the view profile activity
    ActivityLog.objects.create(
        user=request.user,
        activity_type='view_profile',
        activity_description=f'Viewed profile page',
        object_name=request.user.username,
        object_id=request.user.id,
        timestamp=now()
    )

    if request.method == 'POST':
        # Update User model
        request.user.first_name = request.POST.get('first_name', request.user.first_name)
        request.user.last_name = request.POST.get('last_name', request.user.last_name)
        request.user.email = request.POST.get('email', request.user.email)
        request.user.save()

        # Update Profile model
        profile.bio = request.POST.get('bio', profile.bio)
        profile.location = request.POST.get('location', profile.location)
        profile.phone_number=request.POST.get('phone_number', profile.phone_number)
        # If a new profile image is uploaded, update the profile image
        if 'profile_image' in request.FILES:
            profile.profile_image = request.FILES['profile_image']
        profile.save()

        # Log the profile update activity
        ActivityLog.objects.create(
            user=request.user,
            activity_type='update_profile',
            activity_description=f'Updated profile details: first name, last name, email, bio, location.',
            object_name=request.user.username,
            object_id=request.user.id,
            timestamp=now()
        )

        messages.success(request, 'Profile updated successfully!')
        return redirect('profile')  # Redirect to the profile page after saving

    context = {
        'user': request.user,
        'profile': profile,
    }
    return render(request, 'profile.html', context)



@login_required
def home(request):
    notifications = Notification.objects.filter(recipient=request.user).order_by('-timestamp')
    unread_notifications_count = notifications.filter(is_read=False).count()
    
    context = {
        'notifications': notifications,
        'unread_notifications_count': unread_notifications_count,
    }
    return render(request, "home.html",context)

from django.contrib.auth.models import Group

# User Management Views
@login_required
@role_required(['super_admin','admin'])
def user_create(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        email = request.POST.get('email')
        role = request.POST.get('role')  # Get the role from the form

        if password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return redirect('user_create')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return redirect('user_create')
        elif User.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists.')
            return redirect('user_create')            
        else:
            # Create user and set role
            user = User.objects.create_user(username=username, email=email, password=password, role=role)
            
            # Assign the user to the corresponding group
            role_to_group = {
                'super_admin': 'Super Admin',
                'admin': 'Admin',
                'staff': 'Staff',
                'viewer': 'Viewer',
            }
            group_name = role_to_group.get(role)
            if group_name:
                group, created = Group.objects.get_or_create(name=group_name)
                user.groups.add(group)  # Assign user to the role group

            # Log user creation activity
            ActivityLog.objects.create(
                user=request.user,
                activity_type='create_user',
                activity_description=f'Created user {user.username} with role {role}',
                object_name=user.username,
                object_id=user.id,
                timestamp=now()
            )

            messages.success(request, f'User {user.username} created successfully.')
            return redirect('user_list')

    # Render form
    return render(request, 'users/user_create.html')


@login_required
@role_required(['super_admin','admin'])
def user_list(request):
    # Fetch users and categorize them by roles
    users = User.objects.all()
    
    # Categories
    staff_users = users.filter(role='staff')
    viewer_users = users.filter(role='viewer')
    admin_users = users.filter(role='admin')
    super_admin_users = users.filter(role='super_admin')

    # Activity - You can modify the following as per your criteria for activity.
    active_users = users.filter(is_active=True)
    inactive_users = users.filter(is_active=False)

    # User status (active or inactive) counts
    active_user_count = active_users.count()
    inactive_user_count = inactive_users.count()
    
    # Calculate users per category
    staff_count = staff_users.count()
    viewer_count = viewer_users.count()
    admin_count = admin_users.count()
    super_admin_count = super_admin_users.count()
    leader_count=Team.objects.values('leader').distinct().count()
    return render(request, 'users/user_list.html', {
        'users': users,
        'staff_users': staff_users,
        'viewer_users': viewer_users,
        'admin_users': admin_users,
        'super_admin_users': super_admin_users,
        'active_user_count': active_user_count,
        'inactive_user_count': inactive_user_count,
        'staff_count': staff_count,
        'viewer_count': viewer_count,
        'admin_count': admin_count,
        'super_admin_count': super_admin_count,
        'leader_count':leader_count,
    })

@login_required
@role_required(['super_admin', 'admin'])


def user_edit(request, user_id):
    user = get_object_or_404(User, id=user_id)
    permissions = Permission.objects.filter(codename__in=[
        'user_can_upload_file',
        'user_can_create_folder',
        'user_can_delete_file',
        'user_can_rename_file',
    ])
    user_permissions = user.user_permissions.values_list('codename', flat=True)

    if request.method == 'POST':
        # Handle user details form submission
        if 're_page' in request.POST and request.POST['re_page'] == 'user_edit':
            if request.user.id == user.id:
                role = request.POST.get('role')
                if role != user.role:  # Prevent role changes for the current user
                    messages.error(request, "You cannot change your own role.")
                    return redirect('user_edit', user_id=user.id)

            new_username = request.POST.get('username')
            new_email = request.POST.get('email')
            password = request.POST.get('password')
            role = request.POST.get('role')

            # Check role and team constraints
            if role == "viewer" and user.teams.exists():
                messages.error(request, "A user cannot be assigned the 'Viewer' role while part of a team.")
                return redirect('user_edit', user_id=user.id)

            # Validate unique username and email
            if User.objects.filter(username=new_username).exclude(id=user.id).exists():
                messages.error(request, "The username is already taken. Please choose a different username.")
                return redirect('user_edit', user_id=user.id)

            if User.objects.filter(email=new_email).exclude(id=user.id).exists():
                messages.error(request, "The email address is already registered. Please use a different email address.")
                return redirect('user_edit', user_id=user.id)

            # Update user details
            old_username = user.username
            user.username = new_username
            user.email = new_email
            if password:
                user.set_password(password)

            if role in dict(User.ROLE_CHOICES):
                user.role = role

            user.save()  # Triggers `assign_role_group` in the save method

            # Log the activity
            ActivityLog.objects.create(
                user=request.user,
                activity_type='edit_user',
                activity_description=f'Edited user {old_username} (changed username, email, password, role)',
                object_name=user.username,
                object_id=user.id,
                timestamp=now()
            )

            messages.success(request, f'User {user.username} updated successfully.')
            return redirect('user_list')

        # Handle permissions form submission
        elif 'permissions' in request.POST:
            selected_permissions = request.POST.getlist('permissions')

            # Clear existing permissions and add the new ones
            user.user_permissions.clear()
            for perm_codename in selected_permissions:
                perm = Permission.objects.get(codename=perm_codename)
                user.user_permissions.add(perm)

            messages.success(request, f'Permissions updated for {user.username}.')
            return redirect('user_edit', user_id=user.id)

    # Pass current groups, roles, and permissions to the template
    groups = Group.objects.all()
    return render(request, 'users/user_edit.html', {
        'user': user,
        'groups': groups,
        'permissions': permissions,
        'user_permissions': user_permissions,
    })

def toggle_user_status(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        status = request.POST.get('status')
        
        try:
            user = User.objects.get(pk=user_id)
            user.is_active = (status == 'active')
            user.save()
            messages.success(request, f'User is status changed to {status}')
            return redirect('user_detail', user_id)
        except User.DoesNotExist:
            messages.error(request, 'Cannot able to change user status')
            return redirect('user_detail', user_id)
    
    return JsonResponse({'success': False, 'error': 'Invalid request.'}, status=400)

@login_required
@role_required(['super_admin','admin'])
def delete_user(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        print(user_id)
        user = get_object_or_404(User, id=user_id)

        # Check if the user is trying to delete themselves (optional check)
        if user == request.user:
            messages.success(request, 'You cannot Delete your Account by your self')
            return redirect('user_list')

        # Check if the user is assigned to a milestone
        if Milestone.objects.filter(assigned_to=user).exists():
            messages.error(request, 'This user cannot be deleted because they are assigned to a milestone.')
            return redirect('user_list')

        # Check if the user is the leader of any team
        if Team.objects.filter(leader=user).exists():
            messages.error(request, 'This user cannot be deleted because they are a team leader.')
            return redirect('user_list')

        # Log user deletion activity
        ActivityLog.objects.create(
            user=request.user,
            activity_type='delete_user',
            activity_description=f'Deleted user {user.username}',
            object_name=user.username,
            object_id=user.id,
            timestamp=now()
        )

        # Delete the user
        user.delete()

        # Redirect to the user list view after deletion
        messages.success(request, f'User {user.username} has been deleted successfully.')
        return redirect('user_list')

    return HttpResponseForbidden("Invalid request.")

@login_required
@role_required(['super_admin','admin'])
def user_detail(request, user_id):
    user = get_object_or_404(User, id=user_id)

    # Fetch associated data
    teams = Team.objects.filter(members=user)
    milestones = Milestone.objects.filter(assigned_to=user).select_related('project')
    milestone_projects = Project.objects.filter(milestones__assigned_to=user).distinct()
    team_projects = Project.objects.filter(assigned_team__members=user).distinct()
    projects = (milestone_projects | team_projects).distinct()

    # Fetch activity logs for this user
    activities = ActivityLog.objects.filter(user=user).order_by('-timestamp')

    context = {
        'user': user,
        'teams': teams,
        'milestones': milestones,
        'projects': projects,
        'activities': activities,
    }
    return render(request, 'users/user_detail.html', context)


class CustomPasswordResetCompleteView(PasswordResetCompleteView):
    def get(self, *args, **kwargs):
        """Override the GET method to skip default rendering"""
        # Add a success message
        messages.success(self.request, "Your password has been successfully reset. Please log in with your new password.")

        # Redirect to the login page
        return redirect(reverse_lazy('login'))  # Replace 'login' with the correct name of your login URL if different 


def password_reset(request):
    if request.method == "POST":
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            # Now we can safely access cleaned_data
            email = form.cleaned_data.get('email')

            # Check if the email exists in the system
            if User.objects.filter(email=email).exists():
                # Send the reset email
                form.save(request=request)
                messages.success(request, "Password reset link has been sent to your email.")
                return redirect('login')  # Redirect without 'users' namespace if needed
            else:
                # Email doesn't exist, show an error message
                messages.error(request, "The email address provided does not exist.")
        else:
            # Form is not valid
            messages.error(request, "Invalid form submission.")
    else:
        form = PasswordResetForm()

    return render(request, 'users/password_reset.html', {'form': form})

logger = logging.getLogger(__name__)

# Project Views
@login_required
def project_list(request):
    # Retrieve search query and filter from request
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', 'all')

    # Get the current date for calculating remaining days
    today = timezone.now().date()

    # Base queryset
    projects = Project.objects.all()

    # Apply search filter if a query is provided
    if search_query:
        projects = projects.filter(name__icontains=search_query)

    # Apply status filter
    if status_filter == 'completed':
        projects = projects.filter(status='completed')
    elif status_filter == 'in-progress':
        projects = projects.filter(status='ongoing', progress__lt=100)
    elif status_filter == 'pending':
        projects = projects.filter(progress=0)  # Assuming pending means no progress
    elif status_filter == 'deletion-request':
        projects = projects.filter(deletion_requested_by__isnull=False)

    # Annotate additional details for projects
    for project in projects:
        # Ensure start_date and end_date are valid dates
        if project.start_date and project.end_date:
            # Handle valid date calculations for project length and remaining days
            project.length = project.project_length()
            project.remaining_days = project.days_left()

            # Calculate progress
            project.progress = project.calculate_progress()

        else:
            # If either of the dates are invalid, set placeholders
            project.length = 0
            project.remaining_days = 0
            project.progress = 0

        # Get milestones count data if applicable
        milestones = project.milestones.all()
        project.total_milestones = milestones.count()
        project.pending_milestones = milestones.filter(status="pending").count()
        project.in_progress_milestones = milestones.filter(status="in_progress").count()
        project.completed_milestones = milestones.filter(status="completed").count()
        
    context = {
        'projects': projects,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    return render(request, 'projects/project_list.html', context)

def project_timeline(project_id, start_date=None, end_date=None, activity_type=None):
    # Fetch activity logs for the given project, with optional filters
    activities = ActivityLog.objects.filter(object_id=project_id).order_by('timestamp')

    # Apply time filters if provided
    if start_date:
        activities = activities.filter(timestamp__gte=start_date)
    if end_date:
        activities = activities.filter(timestamp__lte=end_date)
    
    # Apply activity_type filter if provided
    if activity_type:
        activities = activities.filter(activity_type=activity_type)

    # Prepare data for plotting
    data = {
        "timestamp": [],
        "activity": [],
        "user": [],
        "details": []
    }
    
    # For no activities case
    if not activities:
        return "No activity logs available for this project."

    # Populate the data dictionary
    for activity in activities:
        data["timestamp"].append(activity.timestamp)
        data["activity"].append(activity.activity_type)
        data["user"].append(activity.user.username)
        data["details"].append(activity.updated_object_data)

    df = pd.DataFrame(data)
     # Define a base height and a per-activity increase
    base_height = 500
    height_per_activity = 50  # Increase height for each activity

    # Calculate the total height by base height + (activities count * height increase per activity)
    total_height = base_height + len(df) * height_per_activity
    def format_details(details):
        if not details:  # Check if details is None or empty
            return "No details available"  # Provide a default message for empty details

        # Check if details is already a dictionary
        if isinstance(details, dict):
            formatted_details = []
            for field, change in details.items():
                old_value = change.get("old", "N/A")  # Default to "N/A" if there's no old value
                new_value = change.get("new", "N/A")  # Default to "N/A" if there's no new value
                formatted_details.append(f"Updated '{field}'")
            
            return '<br>'.join(formatted_details)  # Join each change with a line break

        # If it's a JSON string, load it and then process
        try:
            details_dict = json.loads(details)  # If it's a string, parse it as JSON
            return format_details(details_dict)  # Recursively call the function for a dictionary
        except json.JSONDecodeError:
            return "Invalid details format"  # In case details is not a valid JSON string
    # Artificial timeline: Assign constant time intervals
    df['artificial_timestamp'] = pd.date_range(
        start="2023-01-01",  # Arbitrary start date
        periods=len(df),
        freq='10T'  # Equal spacing of 10 minutes (can be adjusted)
    )

    # Plotting the graph
    fig = go.Figure()

    # Add the vertical line connecting all activities
    fig.add_trace(go.Scatter(
        x=[0] * len(df),  # Keep x=0 for a vertical line
        y=df['artificial_timestamp'],
        mode="lines",  # Line mode
        line=dict(color="black", width=5),  # Vertical line style
        showlegend=False
    ))

    # Add branches with longer horizontal lines and annotations
    for i, row in df.iterrows():
        if i % 2 == 0:
            # Left branch
            fig.add_trace(go.Scatter(
                x=[0, -1],  # From center to the left
                y=[row['artificial_timestamp']] * 2,
                mode="lines+markers",
                line=dict(color="blue", width=1),
                marker=dict(size=10, color="blue", symbol="circle"),
                showlegend=False
            ))

            # Add annotation above the left circle
            formatted_details = format_details(row['details'])  # Format the details field

            fig.add_annotation(
                x=-1,
                y=row['artificial_timestamp'],
                text=f" {str(row['artificial_timestamp'].strftime('%Y-%m-%d %H:%M:%S'))}({row['user']})<br>{formatted_details}",
                showarrow=False,
                font=dict(size=10),
                align="center",
                yshift=20
            )
        else:
            # Right branch with longer horizontal lines
            fig.add_trace(go.Scatter(
                x=[0, 1],  # From center to the right, extended to 3
                y=[row['artificial_timestamp']] * 2,
                mode="lines+markers",  # Add markers
                line=dict(color="green", width=1),
                marker=dict(size=10, color="green", symbol="circle"),  # Circle marker
                showlegend=False
            ))
            # Add annotation above the right circle
            fig.add_annotation(
                x=1,  # Centered on the extended line
                y=row['artificial_timestamp'],
                text=f" {str(row['artificial_timestamp'].strftime('%Y-%m-%d %H:%M:%S'))}({row['user']})<br>{formatted_details}",
                showarrow=False,
                font=dict(size=10),
                align="center",
                yshift=20  # Position above the circle
            )

    # Configure layout to hide the y-axis timestamps and show a clean graph
    fig.update_layout(
        xaxis=dict(
            showgrid=False,
            zeroline=False,
            showticklabels=False  # Hide x-axis ticks
        ),
        yaxis=dict(
            title="",  # No y-axis title
            showgrid=False,  # Hide grid lines
            showticklabels=False  # Hide tick labels (timestamps)
        ),
        height=total_height,
        width=800,
        template="plotly_white",
        margin=dict(l=200, r=200)  # Add more padding for longer branches
    )

    return fig.to_html()  # Return the HTML representation of the plot


from datetime import timedelta

def get_project_activity_history(project_id):
    # Fetch relevant logs
    logs = ActivityLog.objects.filter(
        Q(object_id=project_id) & 
        Q(activity_type__in=['assign_team_to_project', 'create_project'])
    ).order_by('timestamp')

    activity_history = []
    first_team_assignment_time = None

    def format_duration(duration):
        if duration.total_seconds() < 86400:  # Less than a day
            hours, remainder = divmod(duration.seconds, 3600)
            minutes, _ = divmod(remainder, 60)
            return f"{hours}h {minutes}m"
        else:  # More than a day
            days = duration.days
            return f"{days} days"

    for i, log in enumerate(logs):
        if log.activity_type == 'create_project':
            # Extract team details from created_object_data
            created_data = log.created_object_data or {}
            team_name = created_data.get("assigned_team_name", "N/A")

            # Identify the timestamp of the first `assign_team_to_project` log
            for subsequent_log in logs[i + 1:]:
                if subsequent_log.activity_type == 'assign_team_to_project':
                    first_team_assignment_time = subsequent_log.timestamp
                    break

            # Ensure both timestamps are timezone-aware and add 5 seconds to unassigned_at
            unassigned_at = first_team_assignment_time if first_team_assignment_time else timezone.now()
            unassigned_at += timedelta(seconds=5)  # Add 5 seconds

            # Calculate duration
            duration = unassigned_at - log.timestamp

            activity_history.append({
                "activity": "Project Created",
                "assigned_at": log.timestamp,
                "unassigned_at": unassigned_at,
                "team_name": team_name,
                "description": log.activity_description or "N/A",
                "duration": format_duration(duration),
            })
        elif log.activity_type == 'assign_team_to_project':
            # Extract old team name from updated_object_data
            updated_data = log.updated_object_data or {}
            assigned_to_data = updated_data.get("assigned_team", {})
            team_name = assigned_to_data.get("new", "N/A")
            assigned_at = log.timestamp

            # Ensure unassigned_at is timezone-aware and add 5 seconds
            if i + 1 < len(logs) and logs[i + 1].activity_type == 'assign_team_to_project':
                unassigned_at = logs[i + 1].timestamp
            else:
                unassigned_at = timezone.now() + timedelta(seconds=5)  # Add 5 seconds

            # Calculate duration
            duration = unassigned_at - assigned_at

            activity_history.append({
                "activity": "Team Assignment",
                "team_name": team_name,
                "assigned_at": assigned_at,
                "unassigned_at": "Until now" if unassigned_at >= timezone.now() else unassigned_at,
                "duration": format_duration(duration),
            })

    return activity_history

    return activity_history

def project_details(request, project_id):
    team_history=get_project_activity_history(project_id)
    
    project = get_object_or_404(Project, id=project_id)
     # List of activity types
    activity_type_choices = [
        ('cancel_deletion_request', 'Cancel Deletion Request'),
        ('create_project', 'Create Project'),
        ('edit_project', 'Edit Project'),
        ('assign_team_to_project', 'Assign Team to Project'),
    ]
    
    # Get related information
    team = project.assigned_team
    milestones = project.milestones.all()
    completed_milestones = milestones.filter(status='completed')

    # Get today's date
    today = date.today()

    # Overdue milestones: milestones where the end date is before today and their status is not "completed"
    overdue_milestones = milestones.filter(end_date__lt=today, status__in=['pending', 'in_progress'])

    # Count milestone-related information
    total_milestones = milestones.count()
    completed_milestone_count = completed_milestones.count()
    overdue_milestone_count = overdue_milestones.count()

    # Get activity log for the project (for example, filter activity type)
    activity_logs = ActivityLog.objects.filter(object_name=project.name, activity_type="project_team_change").order_by('-timestamp')

    # Get filter parameters from the GET request (optional, defaults to None if not present)
    start_date = request.GET.get('start_date', None)
    end_date = request.GET.get('end_date', None)
    activity_type = request.GET.get('activity_type', None)

    # Convert start_date and end_date to datetime if provided
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

    # Pass filters to project_timeline function
    graph_html = project_timeline(project_id, start_date=start_date, end_date=end_date, activity_type=activity_type)
     # Detect if there are filter parameters in the request
    timeline_active = False
    if 'start_date' in request.GET or 'end_date' in request.GET or 'activity_type' in request.GET:
        timeline_active = True  # If the filter parameters are found, activate the timeline tab
    context = {
        'project': project,
        'team': team,
        'total_milestones': total_milestones,
        'completed_milestone_count': completed_milestone_count,
        'overdue_milestone_count': overdue_milestone_count,
        'activity_logs': activity_logs,
        'milestones': milestones,
        'graph_html': graph_html,
        'timeline_active':timeline_active,
        'activity_type_choices': activity_type_choices,  
        'team_history':team_history
    }

    return render(request, 'projects/project_detail.html', context)


EXTENSION_ICON_MAP = {
    'pdf': 'bi-file-earmark-pdf',
    'jpg': 'bi-file-earmark-image', 
    'jpeg': 'bi-file-earmark-image',  # Add .jpeg to the map
    'png': 'bi-file-earmark-image',
    'doc': 'bi-file-earmark-word',
    'docx': 'bi-file-earmark-word',
    'xlsx': 'bi-file-earmark-spreadsheet',  
    'pptx': 'bi-file-earmark-slides',  
    'txt': 'bi-file-earmark-text',
    'zip': 'bi-file-earmark-zip', 
    'rar': 'bi-file-earmark-zip', 
    'tar': 'bi-file-earmark-zip',
    'gz': 'bi-file-earmark-zip',      
    'json': 'bi-file-earmark-text',  # Add .json extension
    # Add any additional file extensions here as needed
}

def file_manager(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    base_dir = os.path.join(settings.MEDIA_ROOT, 'project_files', f'Project_{project.id}')
    os.makedirs(base_dir, exist_ok=True)

    # Safeguard current_folder to avoid passing NoneType to os.path.join
    current_folder = request.GET.get('folder', '')  # Get folder or default to empty string
    if not current_folder:
        current_folder = ''  # Fallback to root if no folder is specified

    current_dir = os.path.abspath(os.path.join(base_dir, current_folder))
     # Calculate parent folder
    parent_folder_path = os.path.dirname(current_folder)  # This gives the parent folder path
    parent_folder_name = os.path.basename(parent_folder_path) if parent_folder_path else ''  # This gives the parent folder name
    # Breadcrumb Navigation
    breadcrumb = []
    path_accumulator = ""

    # Normalize the path to always use forward slashes
    normalized_folder_path = current_folder.replace("\\", "/")

    for part in normalized_folder_path.split('/'):
        if part:
            path_accumulator = os.path.join(path_accumulator, part).replace("\\", "/")
            breadcrumb.append({"name": part, "path": f"{request.path}?folder={path_accumulator}"})

    # Handle File Upload
    if request.method == "POST" and request.FILES:
        if not request.user.has_perm("gantt_app.user_can_upload_file"):
            print(f"User {request.user.username} does not have upload permission.")
            messages.error(request, "You do not have permission to upload files.")
            return redirect(f'{request.path}?folder={current_folder}')
        uploaded_file = request.FILES.get('file')
        if uploaded_file:
            file_path = os.path.join(current_dir, uploaded_file.name)  # Ensure that file is uploaded into the current folder
            with open(file_path, 'wb+') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)
            messages.success(request, f"File '{uploaded_file.name}' uploaded successfully!")
            return redirect(f'{request.path}?folder={current_folder}')  # Redirect to correct folder

    # Handle Folder Creation
    if request.method == "POST" and "folder_name" in request.POST:
        if not request.user.has_perm("gantt_app.user_can_create_folder"):
            messages.error(request, "You do not have permission to create folders.")
            return redirect(f'{request.path}?folder={current_folder}')
        folder_name = request.POST.get('folder_name', '').strip()
        if folder_name:
            folder_path = os.path.join(current_dir, folder_name)
            os.makedirs(folder_path, exist_ok=True)
            messages.success(request, f"Folder '{folder_name}' created successfully!")
            return redirect(f'{request.path}?folder={current_folder}')

    # Handle Renaming
    if request.method == "POST" and "rename" in request.POST:
        if not request.user.has_perm("gantt_app.user_can_rename_file"):
            messages.error(request, "You do not have permission to rename files.")
            return redirect(f'{request.path}?folder={current_folder}')
        old_name = request.POST.get('old_name')
        new_name = request.POST.get('new_name')
        if old_name and new_name:
            old_path = os.path.join(current_dir, old_name)
            new_path = os.path.join(current_dir, new_name)
            try:
                os.rename(old_path, new_path)
                messages.success(request, f"Renamed '{old_name}' to '{new_name}'!")
            except FileNotFoundError:
                messages.error(request, f"File or folder '{old_name}' not found.")
        else:
            messages.error(request, "Invalid file/folder name(s).")
        return redirect(f'{request.path}?folder={current_folder}')

   # Handle Moving
    if request.method == "POST" and "move" in request.POST:
        file_or_folder = request.POST.get('file_or_folder')
        target_folder = request.POST.get('target_folder')
        
        if file_or_folder and target_folder:
            # If the target folder is the parent folder, adjust the target folder path
            if target_folder == parent_folder_path:
                # Adjust the target folder path to the parent
                target_folder = os.path.dirname(current_folder)  # Remove the last folder from current folder path
            
            old_path = os.path.join(current_dir, file_or_folder)
            new_path = os.path.join(base_dir, target_folder, file_or_folder)
            
            try:
                # Ensure the target folder exists before moving
                if not os.path.exists(os.path.join(base_dir, target_folder)):
                    os.makedirs(os.path.join(base_dir, target_folder))
                
                shutil.move(old_path, new_path)
                messages.success(request, f"Moved '{file_or_folder}' to '{target_folder}'.")
            except FileNotFoundError:
                messages.error(request, f"File or folder '{file_or_folder}' not found.")
            except Exception as e:
                messages.error(request, f"Error while moving: {str(e)}")
        else:
            messages.error(request, "Please provide both file/folder and target folder.")
        return redirect(f'{request.path}?folder={current_folder}')


    # Handle Deleting
    if request.method == "POST" and "delete" in request.POST:
        if not request.user.has_perm("gantt_app.user_can_delete_file"):
            messages.error(request, "You do not have permission to delete files.")
            return redirect(f'{request.path}?folder={current_folder}')
        file_or_folder = request.POST.get('file_or_folder')
        path_to_delete = os.path.join(current_dir, file_or_folder)
        try:
            if os.path.isdir(path_to_delete):
                shutil.rmtree(path_to_delete)
            else:
                os.remove(path_to_delete)
            messages.success(request, f"Deleted '{file_or_folder}'.")
        except FileNotFoundError:
            messages.error(request, f"File or folder '{file_or_folder}' not found.")
        return redirect(f'{request.path}?folder={current_folder}')

    # Fetch directory contents
    try:
        entries = os.listdir(current_dir)
        folders = [{"name": e, "path": os.path.join(current_folder, e)} for e in entries if os.path.isdir(os.path.join(current_dir, e))]
        files = [
            {
                "name": e,
                "path": os.path.join(current_folder, e),
                "extension": e.split('.')[-1].lower() if '.' in e else '', 
                "icon": EXTENSION_ICON_MAP.get(e.split('.')[-1].lower() if '.' in e else '', 'bi bi-file-earmark'),
            }
            for e in entries if os.path.isfile(os.path.join(current_dir, e))
        ]
    except FileNotFoundError:
        raise Http404("Folder not found.")

    return render(request, 'projects/file_manager.html', {
        'project': project,
        'breadcrumb': breadcrumb,
        'folders': folders,
        'files': files,
        'parent_folder_name':parent_folder_name,
        'parent_folder_path':parent_folder_path
    })



@login_required
@role_required(['super_admin', 'admin'])
def project_create(request):
    title = "Create"
    if request.method == 'POST':
        name = request.POST.get('name')
        start_date = parse_date(request.POST.get('start_date'))
        end_date = parse_date(request.POST.get('end_date'))
        can_be_extended = request.POST.get('can_be_extended') == 'on'
        team_id = request.POST.get('team')  # Get the selected team ID
        description = request.POST.get('description')
        
        # Validate the data
        if not name or not start_date or not end_date or not team_id:
            teams = Team.objects.all()  # Fetch teams again in case of error
            return render(request, 'projects/project_form.html', {'teams': teams, 'title': title})

        if start_date > end_date:
            teams = Team.objects.all()  # Fetch teams again in case of error
            return render(request, 'projects/project_form.html', {'teams': teams, 'title': title})

        # Check if the project name already exists
        if Project.objects.filter(name=name).exists():
            messages.error(request, f"'{name}' has already been created.")
            teams = Team.objects.all()  # Fetch teams again in case of error
            return render(request, 'projects/project_form.html', {'teams': teams, 'title': title})

        try:
            team = Team.objects.get(id=team_id)
        except Team.DoesNotExist:
            messages.error(request, "Team does not exist.")
            return render(request, 'projects/project_form.html', {'teams': teams, 'title':title})

        # Create the project (without assigned_team) first
        project = Project.objects.create(
            name=name,
            start_date=start_date,
            end_date=end_date,
            description=description,
           
        )

        # Now, assign the team to the project (foreign key needs the project to be saved first)
        project.assigned_team = team
        project.save()  # Save the project again to reflect the assigned team

        # Log project creation activity
        ActivityLog.objects.create(
            user=request.user,
            activity_type='create_project',
            activity_description=f'Created project {project.name} with team {project.assigned_team.name}',
            object_name=project.name,
            object_id=project.id,
            created_object_data=json.loads(project.to_json()),
            timestamp=now()
           
        )

        return redirect('project_list')

    # If GET request or form validation fails
    teams = Team.objects.all()  # Fetch all available teams
    return render(request, 'projects/project_form.html', {'teams': teams, 'title': title})

@login_required
@role_required(['super_admin','admin'])
def project_edit(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    title = "Edit"
    activity_type="edit_project"
    if request.method == 'POST':
        name = request.POST.get('name')
        start_date = parse_date(request.POST.get('start_date'))
        end_date = parse_date(request.POST.get('end_date'))
        can_be_extended = request.POST.get('can_be_extended') == 'on'
        team_id = request.POST.get('team')
        description = request.POST.get('description')

        # Validate required fields
        if not name or not start_date or not end_date or not team_id:
            messages.error(request, 'Name, start date, end date, and team are required!')
            return render(request, 'projects/project_form.html', {
                'project': project,
                'teams': Team.objects.all(),
                'error': 'Name, start date, end date, and team are required!'
            })

        if start_date > end_date:
            messages.error(request, 'Start date cannot be later than end date!')
            return render(request, 'projects/project_form.html', {
                'project': project,
                'teams': Team.objects.all(),
                'error': 'Start date cannot be later than end date!'
            })

        # Fetch the selected team
        team = Team.objects.get(id=team_id)

        # Store previous details and track changes
        changes = {}
        if project.name != name:
            changes['name'] = {'old': project.name, 'new': name}
            project.name = name

        if project.start_date != start_date:
            changes['start_date'] = {'old': project.start_date, 'new': start_date}
            project.start_date = start_date

        if project.end_date != end_date:
            changes['end_date'] = {'old': project.end_date, 'new': end_date}
            project.end_date = end_date

        if project.can_be_extended != can_be_extended:
            changes['can_be_extended'] = {'old': project.can_be_extended, 'new': can_be_extended}
            project.can_be_extended = can_be_extended

        if project.description != description:
            changes['description'] = {'old': project.description, 'new': description}
            project.description = description

        if project.assigned_team != team:
            activity_type="assign_team_to_project"
            changes['assigned_team'] = {
                'old': project.assigned_team.name if project.assigned_team else None,
                'new': team.name
            }
            project.assigned_team = team

        # If no changes were made
        if not changes:
            messages.info(request, 'No updates were made to the project.')
            return redirect('project_list')

        # Save the project after changes
        project.save()

        # Log the project update activity
        ActivityLog.objects.create(
            user=request.user,
            activity_type=activity_type,
            activity_description=f'Edited project {project.name}.',
            object_name=project.name,
            object_id=project.id,
            timestamp=now(),
            updated_object_data=changes  # Only store the changed fields
        )

        # Success message
        messages.success(request, f'Project updated successfully! Changes: {", ".join(changes.keys())}')
        return redirect('project_list')

    return render(request, 'projects/project_form.html', {
        'project': project,
        'teams': Team.objects.all(),
        'title': title,
        'activity_logs': activity_logs,
    })

@login_required
@role_required(['super_admin','admin'])
def project_delete(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST':
        if not project.deletion_requested_by:
            # Step 1: First admin requests deletion
            project.deletion_requested_by = request.user
            project.save()
            messages.info(request, "Deletion has been requested. A second admin needs to confirm.")
            return redirect('project_list')

        elif project.deletion_requested_by != request.user:
            # Step 2: Second admin confirms deletion
            # Log project deletion activity
            ActivityLog.objects.create(
                user=request.user,
                activity_type='delete_project',
                activity_description=f'Confirmed deletion of project {project.name}',
                object_name=project.name,
                object_id=project.id,
                timestamp=now(),
                # Store all project attributes in JSON format before deletion
                deleted_object_data=json.loads(project.to_json())
            )

            project.delete()
            messages.success(request, "Project has been deleted after confirmation by a second admin.")
            return redirect('project_list')

        else:
            # Prevent the same admin from confirming their own request
            messages.error(request, "You cannot confirm the deletion you initiated.")
            return redirect('project_list')

    return render(request, 'projects/project_confirm_delete.html', {'project': project})

@role_required(['super_admin','admin'])
def cancel_project_deletion(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if request.method == "POST":
        # Ensure only the user who requested deletion can cancel it
        if project.deletion_requested_by == request.user:
            # Clear the deletion request
            project.deletion_requested_by = None
            project.save()

            # Log cancelation activity
            ActivityLog.objects.create(
                user=request.user,
                activity_type="cancel_deletion_request",
                activity_description=f"Canceled deletion request for project {project.name}",
                object_name=project.name,
                object_id=project.id,
                timestamp=now()
            )
            
            messages.success(request, f"Deletion request for '{project.name}' has been canceled.")
        else:
            messages.error(request, "You cannot cancel this deletion request.")          
    return redirect("project_list")
# Task Views


@login_required
def milestone_list(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    milestones = project.milestones.order_by('start_date')  # Order by start_date for sequence
   
    today = now().date()  # Get today's date
    milestone_list_data = []
    active_found = False  # To ensure only one milestone is active
    next_mile={}
    for milestone in milestones:
        # Determine if this milestone is active
        
        is_active = False
        if milestone.status == "in_progress" or milestone.status == "pending":
            if not active_found :
                is_active = True
                next_mile=milestone
                active_found = True  # Mark that the active milestone is found
       
       
        # Prepare status display
        status_display = dict(Milestone.STATUS_CHOICES).get(milestone.status, milestone.status)

        # Calculate remaining days
        days_left = (milestone.end_date - today).days if milestone.end_date and milestone.end_date > today else None

        # Calculate delayed days
        delayed_days = (today - milestone.end_date).days if milestone.end_date and milestone.end_date < today else None

        # Prepare milestone details for the template
        milestone_list_data.append({
            "id": milestone.id,
            "name": milestone.name,
            "description": milestone.description or "No description provided",
            "start_date" : milestone.start_date.strftime('%Y-%m-%d') if milestone.start_date else None,
            "end_date" : milestone.end_date.strftime('%Y-%m-%d') if milestone.end_date else None,
            "status": status_display,
            "days_left": f"{days_left} days" if days_left else "No days",
            "delayed_days": f"{delayed_days} days delayed" if delayed_days else None,
            "assigned_to": [user.username for user in milestone.assigned_to.all()],  # List of usernames
            "is_active": is_active,  # Add active status

        })

    return render(request, "milestones/milestone_list.html", {
        "project": project,
        "milestones": json.dumps(milestone_list_data),  
        "milestoness": milestone_list_data,# Send structured milestone data
        "is_next_active_or_completed": next_mile,

    })




def milestone_create(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if Milestone.objects.filter(project=project_id).exists():
        last_milestone = Milestone.objects.filter(project=project_id).latest("created_at")
    else:
        last_milestone = None  # Or handle it in some way
   
    if request.method == 'POST':
        name = request.POST.get('name')
        start_date = parse_date(request.POST.get('start_date'))
        end_date = parse_date(request.POST.get('end_date'))
        status = "pending"
        description = request.POST.get('description')
        # Validation checks
        if not name or not start_date or not end_date:
            messages.error(request, 'Name, start date, and end date are required!')
            return render(request, 'milestones/milestone_form.html', {
                'project': project,
                'last_milestone':last_milestone,
                'title':"Create"
            })

        if start_date > end_date :
            messages.error(request, 'Start date cannot be later than end date!')
            return render(request, 'milestones/milestone_form.html', {
                'project': project,
                'last_milestone':last_milestone,
                'title':"Create"

            })
        if last_milestone != None:
            if start_date < last_milestone.end_date:
                messages.error(request, 'Start date should be greater than end date of the last Milestone')
                return render(request, 'milestones/milestone_form.html', {
                    'project': project,
                    'last_milestone':last_milestone,
                    'title':"Create"

                })
        # **Additional Validation for Task Dates within Project Duration**
        if start_date < project.start_date or end_date > project.end_date:
            messages.error(request, 'Task should be in a project time scope')
            return render(request, 'milestones/milestone_form.html', {
                'project': project,
                'last_milestone':last_milestone,
                'title':"Create"

            })

        # Check if a milestone with the same name already exists for the current project
        if Milestone.objects.filter(name=name, project=project).exists():
            messages.error(request, f"'{name}' milestone already exists for this project.")
            return render(request, 'milestones/milestone_form.html', {
                'project': project,
                'last_milestone':last_milestone,
                 'title':"Create"
            })

       
        # Create the milestone after validation
        milestone = Milestone.objects.create(
            name=name,
            start_date=start_date,
            end_date=end_date,
            status=status,
            description=description,
            project=project,
        )

       
        # Log the milestone creation activity
        ActivityLog.objects.create(
            user=request.user,
            activity_type='milestone_created',
            activity_description=f'Created milestone "{milestone.name}" for project "{project.name}".',
            object_name=milestone.name,
            object_id=milestone.id,
            timestamp=timezone.now(),
            created_object_data=json.loads(milestone.to_json())
        )

        # Redirect to the milestone list for this project
        messages.success(request, f'Task "{milestone.name}" created successfully.')
        project.update_status_based_on_milestones()  # Update the status based on milestones
        project.save()
        return redirect('milestone_list', project_id=project.id)

    return render(request, 'milestones/milestone_form.html', {
        'project': project,
        'last_milestone':last_milestone,
        'title':"Create"
    })


@project_leader_required
def milestone_edit(request, milestone_id):
    milestone = get_object_or_404(Milestone, id=milestone_id)
    project = get_object_or_404(Project, id=milestone.project.id)
    activity_type = "edit_milestone"
    if request.method == 'POST':
        # Capture original details for comparison
        original_details = {
            "name": milestone.name,
            "start_date": milestone.start_date,
            "end_date": milestone.end_date,
            "description": milestone.description,
        }

        # Fetch updated details from the form
        updated_details = {
            "name": request.POST.get('name'),
            "start_date": parse_date(request.POST.get('start_date')),
            "end_date": parse_date(request.POST.get('end_date')),
            "description": request.POST.get('description'),
        }

       # Detect changes
        # Detect changes
        changes = {}
        for field, original_value in original_details.items():
            updated_value = updated_details[field]

            # Ensure dates are properly converted for comparison
            if field in ["start_date", "end_date"]:
                if isinstance(original_value, (str, type(None))):  
                    original_serialized = original_value  # It's already a string or None
                else:
                    original_serialized = original_value.isoformat()  # Convert date to string
                
                if isinstance(updated_value, (str, type(None))):
                    updated_serialized = updated_value  # Already a string or None
                else:
                    updated_serialized = updated_value.isoformat()  # Convert to string
                
                if original_serialized != updated_serialized:
                    changes[field] = {"old": original_serialized, "new": updated_serialized}

            elif original_value != updated_value:
                changes[field] = {"old": original_value, "new": updated_value}

        # If no changes detected, skip saving
        if not changes:
            messages.info(request, "No changes were detected.")
            return redirect('milestone_list', project_id=project.id)

        # Update only the changed fields
        if "name" in changes:
            if Milestone.objects.filter(name=updated_details["name"], project=project).exists():
                messages.error(request, f"'{updated_details['name']}' milestone already exists for this project.")
                return render(request, 'milestones/milestone_form.html', {
                    'milestone': milestone,
                    'project': project,
                    'title': "Edit"
                })
            milestone.name = updated_details["name"]

        if "start_date" in changes:
            milestone.start_date = parse_date(updated_details["start_date"]) if isinstance(updated_details["start_date"], str) else updated_details["start_date"]

        if "end_date" in changes:
            milestone.end_date = parse_date(updated_details["end_date"]) if isinstance(updated_details["end_date"], str) else updated_details["end_date"]

        if "description" in changes:
            milestone.description = updated_details["description"]
                # Validate required fields
            if not milestone.name or not milestone.start_date or not milestone.end_date:
                return render(request, 'milestones/milestone_form.html', {
                    'error': 'Name, start date, and end date are required!',
                    'milestone': milestone,
                    'project': project,
                    'title':"Edit"
                })

        # Validate date range
        if milestone.start_date > milestone.end_date:
            return render(request, 'milestones/milestone_form.html', {
                'error': 'Start date cannot be later than end date!',
                'milestone': milestone,
                'project': project,
                'title':"Edit"
            })

        # Save the milestone with updated fields
        milestone.save()
        project.update_status_based_on_milestones()
        project.save()

        # Log changes in ActivityLog
        change_description = "; ".join([f"{key}: {values['old']} -> {values['new']}" for key, values in changes.items()])
        ActivityLog.objects.create(
            user=request.user,
            activity_type=activity_type,
            activity_description=f'Edited milestone "{milestone.name}" for project "{project.name}". Changes: {change_description}',
            object_name=milestone.name,
            object_id=milestone.id,
            timestamp=timezone.now(),
            updated_object_data=changes,  # Store only the changed attributes
        )

        messages.success(request, "Task updated successfully.")
        return redirect('milestone_list', project_id=project.id)

    return render(request, 'milestones/milestone_form.html', {
        'milestone': milestone,
        'project': project,
        'title':"Edit"
    })


@project_leader_required
def milestone_delete(request, project_id, milestone_id):
    project = get_object_or_404(Project, id=project_id)
    milestone = get_object_or_404(Milestone, id=milestone_id, project=project)

    if request.method == 'POST':
        # Capture milestone details just before deletion
        deleted_milestone_data = {
            "name": milestone.name,
            "start_date": milestone.start_date,
            "end_date": milestone.end_date,
            "can_be_extended": milestone.can_be_extended,
            "status": milestone.status,
            "description": milestone.description,
            "assigned_to_ids": list(milestone.assigned_to.values_list("id", flat=True))
        }

        # Log the milestone deletion activity with deleted object data
        ActivityLog.objects.create(
            user=request.user,
            activity_type='milestone_deleted',
            activity_description=f'Deleted milestone "{milestone.name}" from project "{project.name}".',
            object_name=milestone.name,
            object_id=milestone.id,
            timestamp=timezone.now(),
            deleted_object_data=deleted_milestone_data  # Store the milestone data to deleted_object_data field
        )

        # Now delete the milestone
        milestone.delete()

        # Success message
        messages.success(request, f'Task "{milestone.name}" deleted successfully.')
        return redirect('milestone_list', project_id=project.id)

    return render(request, 'milestones/milestone_confirm_delete.html', {'milestone': milestone, 'project': project})

@project_leader_required
def milestone_status_update(request, milestone_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            new_status = data.get("status")

            milestone = Milestone.objects.get(id=milestone_id)
            old_status = milestone.status  # Record the old status before updating
            milestone.status = new_status
            milestone.save(update_fields=["status"])
            project = get_object_or_404(Project, id=milestone.project.id)
            project.update_status_based_on_milestones()  # Update the status based on milestones
            project.save()
            # Log milestone status update
            ActivityLog.objects.create(
                user=request.user,
                activity_type='milestone_status_updated',
                activity_description=f'Updated milestone "{milestone.name}" status from "{old_status}" to "{new_status}".',
                object_name=milestone.name,
                object_id=milestone.id,
                timestamp=timezone.now(),
            )

            return JsonResponse({"success": True, "new_status": milestone.get_status_display()})

        except Milestone.DoesNotExist:
            return JsonResponse({"success": False, "error": "Task not found"})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})


def milestone_updates(request, project_id, milestone_id):
    milestone = get_object_or_404(Milestone, id=milestone_id, project__id=project_id)
    updates = TaskUpdate.objects.filter(milestone=milestone).order_by("-date")
    
    # Determine the origin path based on the referrer or explicit request parameter
    if "projects" in request.path:
        origin_path = "projects"
    elif "dashboard" in request.META.get("HTTP_REFERER", ""):  # Use HTTP_REFERER to infer navigation
        origin_path = "dashboard"
    else:
        origin_path = None  # Default to no origin if unclear

    if request.method == "POST":
        progress_detail = request.POST.get("progress_detail")
        issues = request.POST.get("issues")
        status = request.POST.get("status")
        current_date = timezone.localtime().date()  # Ensure comparison against current local date

        # Check if an update already exists for today
        if TaskUpdate.objects.filter(milestone=milestone, date=current_date).exists():
            messages.error(request, "An update for today already exists.")
        else:
            # Create a new TaskUpdate instance
            TaskUpdate.objects.create(
                milestone=milestone,
                updated_by=request.user,
                progress_detail=progress_detail,
                issues=issues,
                date=timezone.now()
            )

            # Update Task status if it has changed
            if status and status != milestone.status:
                milestone.status = status
                milestone.save(update_fields=["status"])
                project = get_object_or_404(Project, id=project_id)
                project.update_status_based_on_milestones()  # Update the status based on milestones
                project.save()
                messages.success(request, "Status updated successfully!")

            messages.success(request, "Daily update and status saved successfully!")
        
        return redirect("milestone_updates", project_id=project_id, milestone_id=milestone_id)

    context = {
        "milestone": milestone,
        "updates": updates,
        "project_id": project_id,
        "origin_path": origin_path,  # Include origin path in context
    }
    return render(request, "milestones/milestone_updates.html", context)


@project_leader_required
def task_create(request, milestone_id):
    milestone = get_object_or_404(Milestone, id=milestone_id)
    users = User.objects.all()

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        description = request.POST.get("description", "").strip()
        assigned_to_ids = request.POST.getlist("assigned_to")
        status = request.POST.get("status", "not_started")
        start_date = parse_date(request.POST.get('start_date'))
        end_date = parse_date(request.POST.get('end_date'))

        # Backend validation: Check if task name already exists
        if Task.objects.filter(name=name, milestone=milestone).exists():
            messages.error(request, "Task name already exists! Please choose a different name.")
            return render(request, "tasks/task_create.html", {"milestone": milestone, "users": users})
        
        # Validate task name is not empty
        if not name:
            messages.error(request, "Task name is required.")
            return render(request, "tasks/task_create.html", {"milestone": milestone, "users": users})

        # Validate start date must be on or after the milestone's start date
        if start_date < milestone.start_date:
            messages.error(request, f"Start date must be on or after the milestone's start date ({milestone.start_date}).")
            return render(request, "tasks/task_create.html", {"milestone": milestone, "users": users})

        # Validate end date must be on or before the milestone's end date
        if end_date > milestone.end_date:
            messages.error(request, f"End date must be on or before the milestone's end date ({milestone.end_date}).")
            return render(request, "tasks/task_create.html", {"milestone": milestone, "users": users})

        # Validate that start date is before the end date
        if start_date >= end_date:
            messages.error(request, "Start date must be earlier than the end date.")
            return render(request, "tasks/task_create.html", {"milestone": milestone, "users": users})

        # Create the Task
        task = Task.objects.create(
            name=name, description=description, start_date=start_date, end_date=end_date, status=status, milestone=milestone
        )
        
        if assigned_to_ids:
            task.assigned_to.set(User.objects.filter(id__in=assigned_to_ids))

        # Add Task Steps
        steps = request.POST.getlist("steps[]")
        for i, step_name in enumerate(steps, start=1):
            if step_name.strip():
                TaskStep.objects.create(task=task, name=f"Step{i}", description=step_name.strip(), order=i)

        messages.success(request, "Task created successfully!")
        return redirect("task_list", milestone_id=milestone_id)

    # Render GET request
    return render(request, "tasks/task_create.html", {"milestone": milestone, "users": users})


def edit_task(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    users = User.objects.all()
    
    if request.method == "POST":
        # Capture the updated form data
        name = request.POST.get("name").strip()
        description = request.POST.get("description", "").strip()
        assigned_to_ids = request.POST.getlist("assigned_to")
        status = request.POST.get("status", "not_started")

        if not name:
            messages.error(request, "Task name is required.")
            return render(request, "tasks/task_create.html", {"task": task, "users": users})

        # Update the Task
        task.name = name
        task.description = description
        task.status = status
        task.save()

        # Update the Task assigned users
        if assigned_to_ids:
            task.assigned_to.set(User.objects.filter(id__in=assigned_to_ids))

        # Update Task Steps (Clear and Recreate)
        task.steps.all().delete()  # Clear existing steps
        steps = request.POST.getlist("steps[]")
        for i, step_name in enumerate(steps, start=1):
            if step_name.strip():
                task.steps.create(name=f"Step{i}", description=step_name.strip(), order=i)

        messages.success(request, "Task updated successfully!")
        return HttpResponseRedirect(reverse("task_list", args=[task.milestone.id]))
    
    return render(request, "tasks/task_create.html", {"task": task, "users": users})


@login_required
def task_list(request, milestone_id):
    milestone = get_object_or_404(Milestone, pk=milestone_id)
    tasks = milestone.tasks.all().order_by('-created_at')
    
    if request.method == 'POST':
        if not request.user.has_perm('projects.manage_milestone', milestone):
            messages.error(request, "You don't have permission for this action")
            return redirect('task_list', milestone_id=milestone_id)
            
        task_ids = request.POST.getlist('task_ids')
        
        if 'delete_tasks' in request.POST:
            Task.objects.filter(id__in=task_ids, milestone=milestone).delete()
            messages.success(request, f'Deleted {len(task_ids)} tasks')
            
        elif 'edit_tasks' in request.POST:
            return redirect('edit_task', task_ids[0])
            
    return render(request, 'tasks/task_list.html', {
        'milestone': milestone,
        'tasks': tasks
    })

def manage_task_steps(request, task_id):
    task = get_object_or_404(Task, pk=task_id)
    steps = task.steps.all().order_by('order')
    status = request.POST.get('status')
    if status:   
        step = get_object_or_404(TaskStep, id=request.POST.get('step_id'))
        if step.issues.filter(resolved=False).exists():
            messages.error(request, "All issues must be resolved before completing this step.")
            return redirect('manage_task_steps', task_id=task_id)
        if request.user in task.assigned_to.all() or request.user.role == "admin":
            step.status = request.POST.get('status')
            print(step.status)
            step.save()
    return render(request, 'tasks/task_steps.html', {
        'task': task,
        'steps': steps,
    })


def bulk_step_actions(request, task_id):
    task = get_object_or_404(Task, pk=task_id)
    if request.user not in task.assigned_to.all() and request.user.role != "admin":
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    step_ids = request.POST.getlist('step_ids')
    action = request.POST.get('bulk_action')
    status = request.POST.get('status')

    if action == 'delete_steps':
        TaskStep.objects.filter(id__in=step_ids, task=task).delete()
        messages.success(request, 'Selected steps deleted successfully')
    elif action == 'edit_steps':
        # Implement multi-edit logic or redirect to edit view
        return edit_task_step(request, step_ids[0],1)
    elif status:
        step_id=request.POST.get('step_id')
        step = get_object_or_404(TaskStep, id=request.POST.get('step_id'))
        if step.issues.filter(resolved=False).exists():
            messages.error(request, "All issues must be resolved before completing this step.")
            return redirect('manage_task_steps', task_id=task_id)
        if request.user in task.assigned_to.all() or request.user.role == "admin":
            step.status = request.POST.get('status')
            step.save()

       
        

    return redirect('manage_task_steps', task_id=task_id)


def handle_issues(request, step_id):
    step = get_object_or_404(TaskStep, pk=step_id)
    if request.user not in step.task.assigned_to.all() and request.user.role != "admin":
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    if 'resolve' in request.POST:
        issue_id = request.POST.get('issue_id')
        issue = get_object_or_404(TaskIssue, pk=issue_id)
        issue.resolved = True
        issue.resolved_at = timezone.now()
        issue.remark = request.POST.get('remark', '')
        issue.save()
        messages.success(request, 'Issue marked as resolved')
    elif 'delete' in request.POST:
        issue_id = request.POST.get('issue_id')
        issue = get_object_or_404(TaskIssue, pk=issue_id)
        issue.delete()
    elif 'unresolve' in request.POST:
        issue_id = request.POST.get('issue_id')
        issue = get_object_or_404(TaskIssue, pk=issue_id)
        issue.resolved = False
        issue.save()
    elif 'create' in request.POST:
        description = request.POST.get('description')
        TaskIssue.objects.create(
            step=step,
            description=description,
            reported_by=request.user
        )
        messages.success(request, 'New issue reported')

    return redirect('manage_task_steps', task_id=step.task.id)

@role_required(['super_admin','admin','staff'])
def edit_task_step(request, step_id,x):
    # Retrieve the step to be edited
    step = get_object_or_404(TaskStep, id=step_id)
    
    if request.method == "POST" and x != 1 :
        # Update only if data is provided
        name = request.POST.get("name")
        description = request.POST.get("description")
        order = request.POST.get("order")
        status = request.POST.get("status")
        
        if name:
            step.name = name
        if description:
            step.description = description
        if order:
            step.order = order
        if status:
            step.status = status
        
        # Save updates
        step.save()
        messages.success(request, f"Step {step.name} updated successfully!")    
    # Render form with existing step details
    return render(request, "tasks/task_step_edit.html", {"step": step})


class MilestonesByProject(APIView):
    def get(self, request, project_id):
        milestones = Milestone.objects.filter(project_id=project_id).values('id', 'name')
        return Response(list(milestones))

class TasksByMilestone(APIView):
    def get(self, request, milestone_id):
        tasks = Task.objects.filter(milestone_id=milestone_id).values('id', 'name')
        return Response(list(tasks))

class StepsByTask(APIView):
    def get(self, request, task_id):
        steps = TaskStep.objects.filter(task_id=task_id).values('id', 'description')
        return Response(list(steps))


@login_required
@role_required(['super_admin','admin'])
def team_create(request):
    search_query = request.GET.get('search', '')  # Get the search query if provided

    # Fetch all users excluding inactive and viewer role users
    users = User.objects.filter(is_active=True).exclude(role__in=['viewer', 'admin', 'super_admin'])
    searched_user = users

    # Filter users based on the search query if applicable
    if search_query:
        searched_user = users.filter(username__icontains=search_query)

    # Prepare team data for each user
    users_with_team_data = []
    for user in searched_user:
        teams_led = Team.objects.filter(leader=user)
        team_count = Team.objects.filter(members=user).count()
        users_with_team_data.append({
            "username": user.username,
            "teams_led": teams_led,
            "team_count": team_count,
        })

    # Pagination
    paginator = Paginator(users_with_team_data, 4)  # Show 4 users per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        name = request.POST.get('name')
        team_leader_id = request.POST.get('team_leader')
        team_members_ids = request.POST.getlist('team_members')

        # Validate required fields
        if not name or not team_leader_id:
            messages.error(request, 'Team name and leader are required.')
            return redirect('team_create')
        
        # Check if the team name already exists
        if Team.objects.filter(name=name).exists():
            messages.error(request, f"'{name}' is already been created.")
            return redirect('team_create')

        try:
            # Get the leader user and validate they are not inactive or a viewer
            leader = get_object_or_404(User.objects.filter(is_active=True).exclude(role='viewer'), id=team_leader_id)

            # Create the team
            team = Team.objects.create(name=name, leader=leader)

            # Add members to the team, excluding inactive or viewer role members
            members = User.objects.filter(id__in=team_members_ids, is_active=True).exclude(role='viewer')
            team.members.set(members)
            team.members.add(leader)  # Ensure leader is also a member

         

            messages.success(request, f'Team {team.name} created successfully.')
            return redirect('team_list')
        except Exception as e:
            messages.error(request, f"Error creating team: {str(e)}")
            return redirect('team_create')

    # GET request
    return render(
        request,
        'teams/team_create.html',
        {
            'users': users,
            'searched_user': searched_user,
            'page_obj': page_obj,
            'search_query': search_query,
        },
    )


@login_required
@role_required(['super_admin','admin'])
def team_list(request):
    # Get the search query from the request
    search_query = request.GET.get('search', '')

    # Filter teams by name if there's a search query
    if search_query:
        teams = Team.objects.filter(name__icontains=search_query)
    else:
        teams = Team.objects.all()

    # Paginate the results
    paginator = Paginator(teams, 5)  # Show 5 teams per page

    # Get the page number from the GET parameter, or default to 1 if not present or invalid
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.get_page(page_number)
    except ValueError:
        raise Http404("Page not found.")  # Handle the case when the page number is invalid

    return render(request, 'teams/team_list.html', {
        'page_obj': page_obj,
        'search_query': search_query,
    })
@login_required
@role_required(['super_admin','admin','staff'])
def add_team_member(request, team_id):
    team = get_object_or_404(Team, id=team_id)

    if request.method == 'POST':
        member_id = request.POST.get('member')
        member = User.objects.get(id=member_id)

        if member in team.members.all():
            messages.error(request, "This user is already a team member.")
        else:
            # Add the member to the team
            team.members.add(member)
            messages.success(request, f"{member.username} has been added to the team.")

            # Create Activity Log for the new member addition
            ActivityLog.objects.create(
                user=request.user,  # The user who is adding the team member
                activity_type='team_member_added',  # Activity type for adding member
                activity_description=f'{member.username} has been added to the team "{team.name}".',  # Description
                object_name=team.name,  # Name of the team
                object_id=team.id,  # ID of the team
                timestamp=timezone.now(),  # Timestamp of the log
            )

        return redirect('team_detail', team_id=team.id)

    # Exclude users already in the team
    available_users = User.objects.exclude(id__in=team.members.all())
    return render(request, 'teams/add_member.html', {'team': team, 'available_users': available_users})

@login_required
@role_required(['super_admin','admin','staff'])
def edit_team(request, team_id):
    team = get_object_or_404(Team, id=team_id)

    if team.leader and team.leader not in team.members.all():
        team.members.add(team.leader)
        team.save()

    all_users = User.objects.filter(is_active=True).exclude(role__in=['viewer', 'admin', 'super_admin'])
    

    if request.method == "POST":
        # Update Team Details
        name = request.POST.get("name")
        description = request.POST.get("description")
        if name:
            old_name = team.name
            team.name = name
            ActivityLog.objects.create(
                user=request.user,
                activity_type='team_updated',
                activity_description=f'Team name changed from "{old_name}" to "{name}".',
                object_name=team.name,
                object_id=team.id,
                timestamp=timezone.now(),
            )
        if description:
            team.description = description

        # Replace Team Leader
        new_leader_id = request.POST.get("new_leader")
        if new_leader_id:
            new_leader = get_object_or_404(User, id=new_leader_id)
            if new_leader in team.members.all():
                old_leader = team.leader
                team.leader = new_leader
                ActivityLog.objects.create(
                    user=request.user,
                    activity_type='team_leader_changed',
                    activity_description=f'Leader changed from {old_leader.username} to {new_leader.username}.',
                    object_name=team.name,
                    object_id=team.id,
                    timestamp=timezone.now(),
                )
            else:
                messages.error(request, f"{new_leader.username} must be a member before becoming the leader.")

       # Add Team Members
        new_member_ids = request.POST.getlist("new_members")  # Match the 'name' in the HTML form
        if new_member_ids:
            for new_member_id in new_member_ids:
                try:
                    # Retrieve the new member and validate
                    new_member = get_object_or_404(User.objects.filter(is_active=True), id=new_member_id)

                    # Add member if not already in the team
                    if new_member not in team.members.all():
                        team.members.add(new_member)
                        ActivityLog.objects.create(
                            user=request.user,
                            activity_type='team_member_added',
                            activity_description=f'{new_member.username} added to the team.',
                            object_name=team.name,
                            object_id=team.id,
                            timestamp=timezone.now(),
                        )
                        messages.success(request, f"{new_member.username} added to the team.")
                    else:
                        messages.info(request, f"{new_member.username} is already in the team.")
                except Exception as e:
                    messages.error(request, f"Error adding member {new_member_id}: {str(e)}")


        # Remove Team Member
        remove_member_id = request.POST.get("remove_member")
        if remove_member_id:
            remove_member = get_object_or_404(User, id=remove_member_id)

            # Prevent removal if the member is the team leader
            if remove_member == team.leader:
                messages.error(request, "Cannot remove the team leader as a member. Replace the leader first.")
            else:
                # Check if the user has any uncompleted milestones assigned to them
                milestones_assigned_to_remove_member = Milestone.objects.filter(assigned_to=remove_member, project__assigned_team=team)
                incomplete_milestones = milestones_assigned_to_remove_member.exclude(status='COMPLETED')

                if incomplete_milestones.exists():
                    messages.error(request, f"Cannot remove {remove_member.username} because they have incomplete milestones.")
                    return redirect("team_list")

                # If all milestones are completed, remove them
                team.members.remove(remove_member)
                messages.success(request, f"{remove_member.username} removed from the team.")

                ActivityLog.objects.create(
                    user=request.user,
                    activity_type='team_member_removed',
                    activity_description=f'{remove_member.username} removed from the team.',
                    object_name=team.name,
                    object_id=team.id,
                    timestamp=timezone.now(),
                )


        team.save()
        return redirect("team_list")

    context = {
        "team": team,
        "all_users": all_users,
        "team_members": team.members.all(),
    }
    return render(request, "teams/edit_team.html", context)


@login_required
@role_required(['super_admin','admin'])
def delete_team(request):
    if request.method == 'POST':
        team_id = request.POST.get('team_id')
       
        team = get_object_or_404(Team, id=team_id)
        
        if Project.objects.filter(assigned_team=team).exclude(status='completed').exists():
            messages.error(request, f"Cannot delete team '{team.name}' because it is assigned to incomplete projects.")
            return redirect('team_list')  # Redirect to the team list view

        # Create activity log for deletion
        ActivityLog.objects.create(
            user=request.user,  # The user who is deleting the team
            activity_type='team_deleted',  # The activity type
            activity_description=f'Team "{team.name}" deleted by {request.user.username}.',  # Description
            object_name=team.name,  # Name of the team
            object_id=team.id,  # ID of the team
            timestamp=now(),  # Timestamp for the log
        )

        # Delete the team
        team.delete()
        messages.success(request, f"Team '{team.name}' deleted successfully.")
        return redirect('team_list')  # Redirect to the team list view after deletion

    return HttpResponseForbidden("Invalid request.")


@login_required
@role_required(['super_admin','admin','staff'])
def remove_team_member(request, team_id, member_id):
    team = get_object_or_404(Team, id=team_id)
    member = get_object_or_404(User, id=member_id)

    if member not in team.members.all():
        messages.error(request, "This user is not a member of the team.")
    else:
        team.members.remove(member)

        # Log the removal action
        ActivityLog.objects.create(
            user=request.user,  # The user who removed the member
            activity_type='team_member_removed',  # Activity type for member removal
            activity_description=f'{member.username} was removed from the team "{team.name}".',
            object_name=team.name,  # Team name
            object_id=team.id,  # Team ID
            timestamp=timezone.now(),  # Timestamp
        )

        messages.success(request, f"{member.username} has been removed from the team.")

    return redirect('team_detail', team_id=team.id)


# View team details
@login_required
@role_required(['super_admin','admin','staff'])
def team_detail(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    return render(request, "teams/team_detail.html", {"team": team})

@login_required
@role_required(['super_admin','admin','staff'])
def replace_team_leader(request, team_id):
    # Get the team and ensure the user is allowed to replace the leader
    team = get_object_or_404(Team, id=team_id)
    current_user = request.user

    # Only allow the current team leader or admin to replace the leader
    if team.leader != current_user and not current_user.is_superuser:
        messages.error(request, "You are not authorized to change the team leader.")
        return redirect('team_list')

    if request.method == 'POST':
        new_leader_id = request.POST.get('new_leader')
        
        # Ensure the new leader is a member of the team
        new_leader = User.objects.get(id=new_leader_id)
        if new_leader not in team.members.all():
            messages.error(request, "The selected member is not a valid team member.")
            return render(request, 'replace_leader.html', {'team': team, 'members': team.members.all()})
        
        # Log activity of changing the leader
        old_leader = team.leader  # Current leader before change

        # Replace the leader
        team.leader = new_leader
        team.save()

        # Create activity log for replacing team leader
        ActivityLog.objects.create(
            user=request.user,  # User who made the change
            activity_type='team_leader_replaced',  # Activity type for leader replacement
            activity_description=f"{old_leader.username} has been replaced by {new_leader.username} as the leader of the team {team.name}.",  # Activity description
            object_name=team.name,  # Team name
            object_id=team.id,  # Team ID
            timestamp=timezone.now(),  # Timestamp of the activity
        )

        messages.success(request, f"Team leader has been updated to {new_leader.username}.")
        return redirect('team_list')

    # GET request: Display the form to replace the team leader
    return render(request, 'teams/replace_leader.html', {'team': team, 'members': team.members.all()})


def activity_logs(request):
    # Fetch all activity logs sorted by timestamp (descending)
    logs = ActivityLog.objects.all().order_by('-timestamp')
    
    # Add any additional logic to filter logs (e.g., by user or project if needed)
    # If filtering by project_id is needed, you can do something like:
    # logs = logs.filter(project__id=request.GET.get('project_id', None))

    context = {
        'logs': logs  # Include the logs in the context
    }
    return render(request, 'activity_logs.html', context)