from django.shortcuts import render, get_object_or_404, redirect
import logging
from django.contrib.auth import get_user_model
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.forms import PasswordResetForm
from django.contrib import messages
from django.utils.dateparse import parse_date
from .models import Task, Project, Team,  ChatRoom, Message, TaskUpdate
from django.contrib.auth.decorators import login_required
User = get_user_model()
from .models import Profile
from .models import Task
from django.http import JsonResponse
import plotly.express as px
import pandas as pd
from django.shortcuts import render
from django.http import HttpResponseRedirect
from django.http import HttpResponseForbidden
import re
from django.core.paginator import Paginator
from django.http import Http404
from datetime import date
from django.db.models import Count, F, ExpressionWrapper, DateField
import os
from django.db.models import Count, Prefetch
from django.utils.timezone import now
from django.utils import timezone
import json
import shutil
from django.contrib.auth.models import Permission
from django.http import JsonResponse
from django.db.models import Count
from .models import ActivityLog
from django.conf import settings
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponseForbidden
from .models import ChatRoom, Message, User ,Notification
from .models import Notification
from .decorators import role_required,project_leader_required,team_member_required
from django.core.files.storage import FileSystemStorage
from itertools import chain
import openpyxl
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go
from django.db.models import Count, Avg
from django.contrib.auth.views import PasswordResetCompleteView
from django.urls import reverse_lazy
from django.db.models import Count, Q


def landing_page(request):
    if request.user.is_staff:  # Check if the user is an admin (staff)
        return admin_landing_page(request)
    else:
        return other_landing_page(request)  # Render for non-admin users
def other_landing_page(request):
        
    return render(request, 'landing_page.html')

def admin_landing_page(request):
    # User role check
    user = request.user
 

    # Statistics for Projects
    total_projects = Project.objects.count()
    ongoing_projects = Project.objects.filter(status='in_progress').count()
    completed_projects = Project.objects.filter(status='completed').count()
    pending_projects = Project.objects.filter(status='pending').count()

    # Statistics for Teams
    total_teams = Team.objects.count()
    teams_with_no_projects = Team.objects.filter(projects__isnull=True).count()
    teams_with_more_than_2_projects = (
    Team.objects.annotate(project_count=Count('projects'))
    .filter(project_count__gt=1)
    .count()
    )

    other_teams = total_teams - teams_with_no_projects - teams_with_more_than_2_projects

    # Statistics for Users
    total_users = User.objects.count()
    users_without_teams = User.objects.filter(teams__isnull=True).count()
    users_without_tasks = User.objects.filter(tasks__isnull=True).count()
    users_in_more_than_2_teams = User.objects.annotate(team_count=Count('teams')).filter(team_count__gt=1).count()
    admin_users= User.objects.filter(role=['admin','super_admin']).count()
    # Context preparation
    context = {
        'total_projects': total_projects,
        'ongoing_projects': ongoing_projects,
        'completed_projects': completed_projects,
        'pending_projects': pending_projects,
        'total_teams': total_teams,
        'teams_with_no_projects': teams_with_no_projects,
        'teams_with_more_than_2_projects': teams_with_more_than_2_projects,
        'other_teams': other_teams,
        'total_users': total_users,
        'users_without_teams': users_without_teams,
        'users_without_tasks': users_without_tasks,
        'users_in_more_than_2_teams': users_in_more_than_2_teams,
        
    }

    return render(request, 'admin/landing_page.html', context)

@role_required(['super_admin','admin','staff'])
def report_dashboard(request):
    projects = Project.objects.all()
    teams = Team.objects.all()

    # Filtering
    project_id = request.GET.get("project")
    team_id = request.GET.get("team")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    filtered_projects = projects
    if project_id:
        filtered_projects = filtered_projects.filter(id=project_id)
    if team_id:
        filtered_projects = filtered_projects.filter(assigned_team__id=team_id)
    if start_date:
        filtered_projects = filtered_projects.filter(start_date__gte=start_date)
    if end_date:
        filtered_projects = filtered_projects.filter(end_date__lte=end_date)
    for project in projects:
            project.length = project.project_length()
            project.remaining_days = project.days_left()
            # Calculate progress
            project.progress = project.calculate_progress()
    context = {
        "projects": projects,
        "teams": teams,
        "filtered_projects": filtered_projects,
    }
    return render(request, "admin/report_dashboard.html", context)

@role_required(['super_admin','admin','staff'])
def generate_filtered_excel(request):
    projects = Project.objects.all()

    # Apply filters as in `report_dashboard`
    project_id = request.GET.get("project")
    team_id = request.GET.get("team")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if project_id:
        projects = projects.filter(id=project_id)
    if team_id:
        projects = projects.filter(assigned_team__id=team_id)
    if start_date:
        projects = projects.filter(start_date__gte=start_date)
    if end_date:
        projects = projects.filter(end_date__lte=end_date)

    # Create Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Filtered Reports"

    # Header
    headers = ["Project Name", "Team Name", "Start Date", "End Date", "Progress"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Data
    for row_num, project in enumerate(projects, 2):
        sheet[f"A{row_num}"] = project.name
        sheet[f"B{row_num}"] = project.assigned_team.name if project.assigned_team else "No Team"
        sheet[f"C{row_num}"] = project.start_date
        sheet[f"D{row_num}"] = project.end_date
        sheet[f"E{row_num}"] = project.progress

    # Response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=filtered_report.xlsx"
    workbook.save(response)
    return response


    # Create a new workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Project Overview"

    # Define the report header
    sheet['A1'] = "Project Name"
    sheet['B1'] = "Start Date"
    sheet['C1'] = "End Date"
    sheet['D1'] = "Status"
    sheet['E1'] = "Total Tasks"
    sheet['F1'] = "Completed Tasks"
    sheet['G1'] = "Pending Tasks"
    sheet['H1'] = "Progress (%)"
    sheet['I1'] = "Assigned Team"  # New header for assigned team

    # Fetch all projects
    projects = Project.objects.all()
    row = 2  # Start from the second row for data

    for project in projects:
        tasks = project.tasks.all()

        # Calculate statistics related to tasks
        total_tasks = tasks.count()
        completed_tasks = tasks.filter(status='completed').count()
        pending_tasks = tasks.filter(status='pending').count()
        progress = project.calculate_progress()
        assigned_team = project.assigned_team.name if project.assigned_team else "N/A"  # Fetch the team name

        # Write data to excel
        sheet[f'A{row}'] = project.name
        sheet[f'B{row}'] = project.start_date
        sheet[f'C{row}'] = project.end_date
        sheet[f'D{row}'] = project.status
        sheet[f'E{row}'] = total_tasks
        sheet[f'F{row}'] = completed_tasks
        sheet[f'G{row}'] = pending_tasks
        sheet[f'H{row}'] = progress
        sheet[f'I{row}'] = assigned_team  # Add the assigned team name

        row += 1  # Move to the next row

    # Generate response for Excel download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=project_overview.xlsx'
    workbook.save(response)  # Write the Excel file to the response
    return response

@login_required
def dashboard(request):
    projects = Project.objects.all()
    teams = Team.objects.all()

    # Determine the active tab based on the GET parameters
    active_tab = "project_status"  # Default tab
    if request.GET.get('project_id'):
        active_tab = "tasks_by_status"
    elif request.GET.get('team_id'):
        active_tab = "team_contribution"

    # Get the selected project and generate the chart
    selected_project_id = request.GET.get('project_id')
    project_tasks_chart = tasks_status_chart(project_id=selected_project_id) if selected_project_id else tasks_status_chart()

    # Get the selected team and generate the chart
    selected_team_id = request.GET.get('team_id')
    team_contrib_chart = team_contribution_chart(team_id=selected_team_id) if selected_team_id else team_contribution_chart()
    
    # Context for the template
    context = {
        'teams': teams,
        'team_contrib_chart': team_contrib_chart,
        'projects': projects,
        'project_tasks_chart': project_tasks_chart,
        "project_status_chart": project_status_chart(),
        "tasks_status_chart": tasks_status_chart(),
        "team_contribution_chart": team_contribution_chart(),
        "project_progress_chart": project_progress_chart(),
        "remaining_days_chart": remaining_days_chart(),
        "delayed_projects_chart": delayed_projects_chart(),
        "effort_accuracy_chart": effort_accuracy_chart(),
        "tasks_completed_chart": tasks_completed_chart(),
        "team_task_allocation_chart": team_task_allocation_chart(),
        "active_tab": active_tab,  # Pass the active tab to the template
    }
    return render(request, "dashboard.html", context)



def project_status_chart():
    status_counts = Project.objects.values("status").annotate(count=Count("status"))
    labels = [entry["status"].replace("_", " ").capitalize() for entry in status_counts]
    values = [entry["count"] for entry in status_counts]

    fig = go.Figure(
        go.Pie(
            labels=labels,
            values=values,
            hole=0.3,
            textinfo="label+value",  # Display the label and the actual count
            textposition="inside"
        )
    )
    fig.update_layout(title="Project Status Distribution")
    return fig.to_html()

def tasks_status_chart(project_id=None):
    # Filter tasks by project_id if it's provided, else all tasks
    if project_id:
        task_counts = Task.objects.filter(project_id=project_id).values("status").annotate(count=Count("status"))
    else:
        task_counts = Task.objects.values("status").annotate(count=Count("status"))

    # Prepare data for the chart
    statuses = [entry["status"].capitalize() for entry in task_counts]
    counts = [entry["count"] for entry in task_counts]

    fig = go.Figure(
        go.Bar(
            x=statuses,
            y=counts,
            marker_color=["blue", "orange", "green"],  # You can add more or modify colors as needed
        )
    )
    fig.update_layout(
        title="Task Status Distribution", 
        xaxis_title="Status", 
        yaxis_title="Count"
    )
    
    return fig.to_html()

def team_contribution_chart(team_id=None):
    if team_id:
        team = Team.objects.get(id=team_id)
        members = team.members.all()
        member_task_counts = []
        total_tasks = Task.objects.filter(project__assigned_team=team).count()

        for member in members:
            task_count = Task.objects.filter(assigned_to=member).count()
            percentage = (task_count / total_tasks) * 100 if total_tasks else 0
            member_task_counts.append({
                'username': member.username,
                'task_count': task_count,
                'percentage': round(percentage, 2)
            })

        # Extract data for the chart
        usernames = [entry['username'] for entry in member_task_counts]
        task_counts = [entry['task_count'] for entry in member_task_counts]
        percentages = [entry['percentage'] for entry in member_task_counts]

        # Generate Bar Chart with Task Count and Percentage
        fig = go.Figure()

        # Add bars for task count
        fig.add_trace(go.Bar(
            x=usernames,
            y=task_counts,
            name='Task Count',
            text=task_counts,  # Show task count on the bars
            textposition='auto',
            hoverinfo='x+y',  # Show task count on hover
            marker=dict(color='rgb(55, 83, 109)')
        ))


        fig.update_layout(
            barmode='group',  # Group the bars side by side
            title="Team Members Task Contribution",
            xaxis_title="Team Member",
            yaxis_title="Task Count ",
            yaxis=dict(
                title="Task Count",
                range=[0, max(task_counts )]  # Adjust y-axis range
            )
        )

    else:
        fig = go.Figure()

    return fig.to_html()


 
def project_progress_chart():
    projects = Project.objects.all()
    names = [project.name for project in projects]
    progress_values = [project.calculate_progress()for project in projects]

    fig = go.Figure(go.Bar(x=names, y=progress_values, text=progress_values, textposition="auto"))
    fig.update_layout(
        title="Project Progress Status",
        xaxis_title="Project",
        yaxis_title="Progress (%)",
        yaxis=dict(range=[0, 100]),  # Limit to 0-100%
    )
    return fig.to_html()
def remaining_days_chart():
    projects = Project.objects.filter(end_date__gte=date.today())
    names = [project.name for project in projects]
    remaining_days = [project.days_left() for project in projects]

    fig = go.Figure(go.Bar(x=names, y=remaining_days, text=remaining_days, textposition="auto"))
    fig.update_layout(
        title="Remaining Days for Projects",
        xaxis_title="Project",
        yaxis_title="Days Left",
    )
    return fig.to_html()
def delayed_projects_chart():
    delayed_projects = Project.objects.filter(end_date__lt=date.today(), status__in=["pending", "in_progress"])
    names = [project.name for project in delayed_projects]
    delays = [abs(project.days_left()) for project in delayed_projects]

    fig = go.Figure(go.Bar(x=names, y=delays, text=delays, textposition="auto", marker_color="red"))
    fig.update_layout(
        title="Delayed Projects",
        xaxis_title="Project",
        yaxis_title="Days Delayed",
    )
    return fig.to_html()
def effort_accuracy_chart():
    tasks = Task.objects.filter(effort_estimation__isnull=False, actual_effort__isnull=False)
    task_names = [task.name for task in tasks]
    estimated_effort = [task.effort_estimation for task in tasks]
    actual_effort = [task.actual_effort for task in tasks]

    fig = go.Figure()
    fig.add_trace(go.Bar(x=task_names, y=estimated_effort, name="Estimated Effort", marker_color="blue"))
    fig.add_trace(go.Bar(x=task_names, y=actual_effort, name="Actual Effort", marker_color="orange"))
    fig.update_layout(
        title="Effort Accuracy",
        xaxis_title="Tasks",
        yaxis_title="Hours",
        barmode="group",
    )
    return fig.to_html()
def tasks_completed_chart():
    teams = Team.objects.all()
    team_names = []
    completion_rates = []

    for team in teams:
        # For each team, find all the tasks from the projects assigned to that team
        projects = Project.objects.filter(assigned_team=team)
        total_tasks = 0
        completed_tasks = 0

        # Iterate over projects assigned to the team and count completed tasks
        for project in projects:
            tasks = project.tasks.all()  # Fetch tasks related to the project
            total_tasks += tasks.count()  # Count the total number of tasks for the project
            completed_tasks += tasks.filter(status="completed").count()  # Count completed tasks

        if total_tasks > 0:
            team_names.append(team.name)
            completion_rates.append(round((completed_tasks / total_tasks) * 100, 2))

    fig = go.Figure(go.Bar(x=team_names, y=completion_rates, text=completion_rates, textposition="auto"))
    fig.update_layout(
        title="Task Completion Percentage by Team",
        xaxis_title="Team",
        yaxis_title="Completion Rate (%)",
        yaxis=dict(range=[0, 100]),
    )
    return fig.to_html()

def team_task_allocation_chart():
    teams = Team.objects.all()
    team_names = [team.name for team in teams]
    task_counts = []

    for team in teams:
        # For each team, find all tasks in projects assigned to that team
        projects = Project.objects.filter(assigned_team=team)
        total_tasks = 0
        for project in projects:
            tasks = project.tasks.all()  # Get tasks related to the project
            total_tasks += tasks.count()  # Count tasks assigned to team in the project

        task_counts.append(total_tasks)

    fig = go.Figure(go.Pie(labels=team_names, values=task_counts, hole=0.3))
    fig.update_layout(title="Task Allocation by Team")
    return fig.to_html()

@login_required
def user_dashboard(request):
    user = request.user

    # Get the user's teams
    user_teams = user.teams.all()

    # Get all projects associated with the user via team assignment
    relevant_projects = Project.objects.filter(assigned_team__in=user_teams).distinct()

    # Fetch the tasks
    assigned_tasks = Task.objects.filter(assigned_to=user)
    team_tasks = Task.objects.filter(assigned_group__in=user_teams).distinct()

    # Combine tasks
    relevant_tasks = list(chain(assigned_tasks, team_tasks))

    # Create a dictionary to group detailed task data by project
    project_tasks = {}
    today = now().date()
    for task in relevant_tasks:
        if task.project not in project_tasks:
            project_tasks[task.project] = []

        # Prepare task details
        status_display = dict(Task.STATUS_CHOICES).get(task.status, task.status)
        days_left = (task.end_date - today).days if task.end_date and task.end_date > today else None
        delayed_days = (today - task.end_date).days if task.end_date and task.end_date < today else None

        project_tasks[task.project].append({
            "id": task.id,
            "name": task.name,
            "description": task.description or "No description provided",
            "start_date": task.start_date,
            "end_date": task.end_date,
            'status': status_display,
            "days_left": f"{days_left} days " if days_left else "No days",
            "delayed_days": f"{delayed_days} days delayed" if delayed_days else None,
            "assigned_to": [user.username for user in task.assigned_to.all()],  # List of usernames
        })

    context = {
        'user': user,
        'projects': relevant_projects,
        'project_tasks': project_tasks,
    }
    return render(request, 'users/users_dashboard.html', context)
@role_required(['super_admin','admin','staff'])
def notifications_view(request):
    notifications = request.user.notifications.filter(is_read=False).order_by('-timestamp')
    return render(request, 'notifications.html', {'notifications': notifications})


@role_required(['super_admin','admin','staff'])
def mark_notification_read(request, notification_id):
    if request.method == 'POST':
        try:
            notification = Notification.objects.get(id=notification_id, recipient=request.user)
            notification.is_read = True
            notification.save()
            return redirect(f"{request.META.get('HTTP_REFERER', '/')}")
        except Notification.DoesNotExist:
            return redirect(f"{request.META.get('HTTP_REFERER', '/')}")
    return redirect(f"{request.META.get('HTTP_REFERER', '/')}")


@role_required(['super_admin', 'admin', 'staff'])
def send_message(request, room_id):
    # Get the chat room
    chat_room = get_object_or_404(ChatRoom, id=room_id)
    # Check if the user is a member of the project
    project = chat_room.project.assigned_team
    
    if request.user not in project.members.all():
        messages.error(request, 'You are not a member of this project.')
        return redirect('chat_room_detail', room_id=room_id)

    if request.method == 'POST':
        content = request.POST.get('content')
        reply_to_id = request.POST.get('reply_to')
        file = request.FILES.get('file')

        # Check if the content is not empty or there is no file uploaded
        if not content.strip() and not file:
            messages.error(request, 'Please write a message or upload a file.')
            return redirect('chat_room_detail', room_id=room_id)

        # Handle Mentions (extract usernames from @username syntax)
        mentioned_usernames = re.findall(r'@(\w+)', content)
        mentions = User.objects.filter(username__in=mentioned_usernames)

        # Create a new message object
        message = Message(
            user=request.user,
            room=chat_room,
            content=content,
            timestamp=now(),
        )

        # If the message is a reply, link the reply
        if reply_to_id:
            try:
                reply_to_message = Message.objects.get(id=reply_to_id)
                message.reply_to = reply_to_message
            except Message.DoesNotExist:
                messages.error(request, 'The message you are replying to does not exist.')
                return redirect('chat_room_detail', room_id=room_id)

        # Handle file upload (if present)
        if file:
            try:
                message.file = file
            except ValidationError as e:
                messages.error(request, f'File upload error: {str(e)}')
                return redirect('chat_room_detail', room_id=room_id)

        # Save the message
        message.save()

        # Log this activity
        ActivityLog.objects.create(
            user=request.user,  # The user who sent the message
            activity_type='message_sent',  # Define the activity type
            activity_description=f'Message sent in chat room "{chat_room.name}".',  # Description of the action
            object_name=chat_room.name,  # Chat room name where the message was sent
            object_id=chat_room.id,  # ID of the chat room
            timestamp=now(),  # Timestamp of the action
        )

        # Notify mentioned users
        for user in mentions:
            Notification.objects.create(
                recipient=user,
                message=f"You were mentioned in a chat by {request.user.username}."
            )

        # Redirect back to the chat room with a success message
        messages.success(request, 'Message sent successfully.')
        return redirect('chat_room_detail', room_id=room_id)

    # If it's not a POST request, redirect to the chat room
    return redirect('chat_room_detail', room_id=room_id)


@role_required(['super_admin','admin','staff'])
def edit_message(request, room_id, message_id):
    # Get the room and message to edit
    chat_room = get_object_or_404(ChatRoom, id=room_id)
    message = get_object_or_404(Message, id=message_id)

    # Check if the user is allowed to edit the message (creator or admin)
    if message.user == request.user or request.user.is_staff:
        if request.method == 'POST':
            # Get new content and file from the form
            new_content = request.POST.get('content')
            new_file = request.FILES.get('file')

            # Update message content
            old_content = message.content  # Save original content to compare later
            message.content = new_content
            if new_file:
                message.file = new_file  # Update the file if a new one is provided

            message.save()

            # Log this activity for editing the message
            ActivityLog.objects.create(
                user=request.user,
                activity_type='message_edited',  # Define activity type
                activity_description=f'Message edited in chat room "{chat_room.name}". Old content: "{old_content}".',
                object_name=chat_room.name,
                object_id=chat_room.id,
                timestamp=now(),
            )

            messages.success(request, 'Message updated successfully!')
            return redirect('chat_room_detail', room_id=room_id)
        else:
            messages.error(request, 'Invalid request method.')
    else:
        messages.error(request, 'You are not authorized to edit this message.')

    return redirect('chat_room_detail', room_id=room_id)


@role_required(['super_admin','admin','staff'])
def delete_message(request, message_id):
    # Get the message object using the message_id
    message = get_object_or_404(Message, id=message_id)

    # Check if the user has permission to delete the message (optional)
    if message.user != request.user and not request.user.is_staff:
        return HttpResponse("You do not have permission to delete this message.", status=403)

    # Log the activity for deleting the message
    ActivityLog.objects.create(
        user=request.user,
        activity_type='message_deleted',  # Define activity type
        activity_description=f'Message deleted in chat room "{message.room.name}".',
        object_name=message.room.name,  # The chat room where the message was deleted
        object_id=message.room.id,  # ID of the chat room
        timestamp=now(),
    )

    # Delete the message
    room_id = message.room.id
    message.delete()

    # Redirect to the chat room or message list after deletion
    return redirect('chat_room_detail', room_id=room_id)



def chat_room_detail(request, room_id):
    chat_room = get_object_or_404(ChatRoom, project=room_id)
    
    messages_list = Message.objects.filter(room=chat_room).order_by('timestamp')
    project = get_object_or_404(Project, id=room_id)
 
    team = chat_room.project.assigned_team
    team_members = team.members.select_related('profile') if team else []
    
    # Log activity when entering the chat room (viewing)
    ActivityLog.objects.create(
        user=request.user,
        activity_type='view_chat_room',
        activity_description=f'Viewed chat room "{chat_room.name}"',
        object_name=chat_room.name,
        object_id=chat_room.id,
        timestamp=now()
    )

    # Handling message edits
    if request.method == 'POST' and 'edit_message_id' in request.POST:
        message_id = request.POST.get('edit_message_id')
        new_content = request.POST.get('new_content')
        new_file = request.FILES.get('new_file')

        # Get the message to be edited
        message = get_object_or_404(Message, id=message_id)

        # Ensure the user editing the message is the message's creator or an admin
        if message.user == request.user or request.user.is_staff:
            # Log the activity before the update to track the content
            old_content = message.content  # Save old content before it is changed
            
            # Update the content
            message.content = new_content
            if new_file:
                message.file = new_file  # Update the file if provided
            message.save()

            # Log the edit activity
            ActivityLog.objects.create(
                user=request.user,
                activity_type='edit_message',
                activity_description=f'Edited message in chat room "{chat_room.name}". Old content: "{old_content}".',
                object_name=chat_room.name,
                object_id=chat_room.id,
                timestamp=now()
            )

            # Show success message and redirect
            messages.success(request, 'Message updated successfully!')
            return redirect('chat_room_detail', room_id=room_id)

        # If the user doesn't have permission to edit, show an error
        messages.error(request, 'You do not have permission to edit this message.')
        return redirect('chat_room_detail', room_id=room_id)

    # Return the rendered template with necessary context
    return render(request, 'chats/chat_room_detail.html', {
        'chat_room': chat_room,
        'messagess': messages_list,
        'team_members': team_members,
        'user': request.user,
        'project': project,
    })


def chatroom(request,room_id):
    return render(request,'chats/chatroom.html')


def gantt_chart(request, project_id):
    # Fetch the project and its tasks
    project = get_object_or_404(Project, id=project_id)
    tasks = project.tasks.all()  # Get all tasks related to the project

    # Check if there are no tasks
    if not tasks:
        return render(request, 'index.html', {
            'chart_html': None,
            'project': project,
            'message': "No tasks available for this project."
        })
    
    # Prepare data for the Gantt chart
    data = []
    for task in tasks:
        # Handle many-to-many assignment
        assigned_users = ', '.join([user.username for user in task.assigned_to.all()]) if task.assigned_to.exists() else 'Unassigned'
        
        data.append({
            'Task': task.name,
            'Start': task.start_date,
            'Finish': task.end_date,
            'Assigned To': assigned_users,
        })
    
    # Convert data to a pandas DataFrame
    df = pd.DataFrame(data)
    
    # Create the Gantt chart
    fig = px.timeline(df, x_start="Start", x_end="Finish", y="Task", color="Assigned To", title=f"{project.name}")
    fig.update_yaxes(categoryorder="total ascending")  # Sort tasks by start date
    
    # Convert the chart to HTML
    chart_html = fig.to_html(full_html=False)
    
    # Render the template with the Gantt chart
    return render(request, 'index.html', {'chart_html': chart_html, 'project': project})


def custom_login(request):
    # Check if 'next' is present in the URL parameters
    next_url = request.GET.get('next', None)

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        try:
            # Try to find the user manually first
            user = User.objects.get(username=username)

            # Check if user is inactive before authenticating
            if not user.is_active:
                # Log the failed login attempt due to inactive user status
                ActivityLog.objects.create(
                    user=user,
                    activity_type='failed_login',
                    activity_description=f'Login attempt blocked due to inactive status for username {username}',
                    object_name=username,
                    object_id=user.id,
                    timestamp=now()
                )
                messages.error(request, 'Your account is inactive. Please contact an administrator.')
                return redirect('login')  # Prevent login for inactive users

            # Proceed with authentication if user is active
          
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                # Log successful login activity
                ActivityLog.objects.create(
                    user=user,
                    activity_type='login',
                    activity_description='Successful login',
                    object_name=user.username,
                    object_id=user.id,
                    timestamp=now()
                )

                # Redirect based on 'next' or default to home page
                return redirect(next_url or 'home')

            else:
                # Handle failed login due to invalid credentials
                if user is not None:
                    ActivityLog.objects.create(
                        user=None,  # No user associated with invalid credentials
                        activity_type='failed_login',
                        activity_description=f'Failed login attempt for username {username} - Invalid credentials',
                        object_name=username,
                        object_id=None,
                        timestamp=now()
                    )
                messages.error(request, 'Invalid username or password')
                return redirect('login')

        except User.DoesNotExist:
            messages.error(request, 'Invalid username or password')
            return redirect('login')

    return render(request, 'users/login.html')
  
@login_required
def profile(request):
    
    # Get or create the profile for the logged-in user
    profile, created = Profile.objects.get_or_create(user=request.user)

    # Log the view profile activity
    ActivityLog.objects.create(
        user=request.user,
        activity_type='view_profile',
        activity_description=f'Viewed profile page',
        object_name=request.user.username,
        object_id=request.user.id,
        timestamp=now()
    )

    if request.method == 'POST':
        # Update User model
        request.user.first_name = request.POST.get('first_name', request.user.first_name)
        request.user.last_name = request.POST.get('last_name', request.user.last_name)
        request.user.email = request.POST.get('email', request.user.email)
        request.user.save()

        # Update Profile model
        profile.bio = request.POST.get('bio', profile.bio)
        profile.location = request.POST.get('location', profile.location)
        profile.phone_number=request.POST.get('phone_number', profile.phone_number)
        # If a new profile image is uploaded, update the profile image
        if 'profile_image' in request.FILES:
            profile.profile_image = request.FILES['profile_image']
        profile.save()

        # Log the profile update activity
        ActivityLog.objects.create(
            user=request.user,
            activity_type='update_profile',
            activity_description=f'Updated profile details: first name, last name, email, bio, location.',
            object_name=request.user.username,
            object_id=request.user.id,
            timestamp=now()
        )

        messages.success(request, 'Profile updated successfully!')
        return redirect('profile')  # Redirect to the profile page after saving

    context = {
        'user': request.user,
        'profile': profile,
    }
    return render(request, 'profile.html', context)



@login_required
def home(request):
    notifications = Notification.objects.filter(recipient=request.user).order_by('-timestamp')
    unread_notifications_count = notifications.filter(is_read=False).count()
    
    context = {
        'notifications': notifications,
        'unread_notifications_count': unread_notifications_count,
    }
    return render(request, "home.html",context)

from django.contrib.auth.models import Group

# User Management Views
@login_required
@role_required(['super_admin','admin'])
def user_create(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        email = request.POST.get('email')
        role = request.POST.get('role')  # Get the role from the form

        if password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return redirect('user_create')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return redirect('user_create')
        elif User.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists.')
            return redirect('user_create')            
        else:
            # Create user and set role
            user = User.objects.create_user(username=username, email=email, password=password, role=role)
            
            # Assign the user to the corresponding group
            role_to_group = {
                'super_admin': 'Super Admin',
                'admin': 'Admin',
                'staff': 'Staff',
                'viewer': 'Viewer',
            }
            group_name = role_to_group.get(role)
            if group_name:
                group, created = Group.objects.get_or_create(name=group_name)
                user.groups.add(group)  # Assign user to the role group

            # Log user creation activity
            ActivityLog.objects.create(
                user=request.user,
                activity_type='create_user',
                activity_description=f'Created user {user.username} with role {role}',
                object_name=user.username,
                object_id=user.id,
                timestamp=now()
            )

            messages.success(request, f'User {user.username} created successfully.')
            return redirect('user_list')

    # Render form
    return render(request, 'users/user_create.html')


@login_required
@role_required(['super_admin','admin'])
def user_list(request):
    # Fetch users and categorize them by roles
    users = User.objects.all()
    
    # Categories
    staff_users = users.filter(role='staff')
    viewer_users = users.filter(role='viewer')
    admin_users = users.filter(role='admin')
    super_admin_users = users.filter(role='super_admin')

    # Activity - You can modify the following as per your criteria for activity.
    active_users = users.filter(is_active=True)
    inactive_users = users.filter(is_active=False)

    # User status (active or inactive) counts
    active_user_count = active_users.count()
    inactive_user_count = inactive_users.count()
    
    # Calculate users per category
    staff_count = staff_users.count()
    viewer_count = viewer_users.count()
    admin_count = admin_users.count()
    super_admin_count = super_admin_users.count()

    return render(request, 'users/user_list.html', {
        'users': users,
        'staff_users': staff_users,
        'viewer_users': viewer_users,
        'admin_users': admin_users,
        'super_admin_users': super_admin_users,
        'active_user_count': active_user_count,
        'inactive_user_count': inactive_user_count,
        'staff_count': staff_count,
        'viewer_count': viewer_count,
        'admin_count': admin_count,
        'super_admin_count': super_admin_count,
    })

@login_required
@role_required(['super_admin', 'admin'])


def user_edit(request, user_id):
    user = get_object_or_404(User, id=user_id)
    permissions = Permission.objects.filter(codename__in=[
        'user_can_upload_file',
        'user_can_create_folder',
        'user_can_delete_file',
        'user_can_rename_file',
    ])
    user_permissions = user.user_permissions.values_list('codename', flat=True)

    if request.method == 'POST':
        # Handle user details form submission
        if 're_page' in request.POST and request.POST['re_page'] == 'user_edit':
            if request.user.id == user.id:
                role = request.POST.get('role')
                if role != user.role:  # Prevent role changes for the current user
                    messages.error(request, "You cannot change your own role.")
                    return redirect('user_edit', user_id=user.id)

            new_username = request.POST.get('username')
            new_email = request.POST.get('email')
            password = request.POST.get('password')
            role = request.POST.get('role')

            # Check role and team constraints
            if role == "viewer" and user.teams.exists():
                messages.error(request, "A user cannot be assigned the 'Viewer' role while part of a team.")
                return redirect('user_edit', user_id=user.id)

            # Validate unique username and email
            if User.objects.filter(username=new_username).exclude(id=user.id).exists():
                messages.error(request, "The username is already taken. Please choose a different username.")
                return redirect('user_edit', user_id=user.id)

            if User.objects.filter(email=new_email).exclude(id=user.id).exists():
                messages.error(request, "The email address is already registered. Please use a different email address.")
                return redirect('user_edit', user_id=user.id)

            # Update user details
            old_username = user.username
            user.username = new_username
            user.email = new_email
            if password:
                user.set_password(password)

            if role in dict(User.ROLE_CHOICES):
                user.role = role

            user.save()  # Triggers `assign_role_group` in the save method

            # Log the activity
            ActivityLog.objects.create(
                user=request.user,
                activity_type='edit_user',
                activity_description=f'Edited user {old_username} (changed username, email, password, role)',
                object_name=user.username,
                object_id=user.id,
                timestamp=now()
            )

            messages.success(request, f'User {user.username} updated successfully.')
            return redirect('user_list')

        # Handle permissions form submission
        elif 'permissions' in request.POST:
            selected_permissions = request.POST.getlist('permissions')

            # Clear existing permissions and add the new ones
            user.user_permissions.clear()
            for perm_codename in selected_permissions:
                perm = Permission.objects.get(codename=perm_codename)
                user.user_permissions.add(perm)

            messages.success(request, f'Permissions updated for {user.username}.')
            return redirect('user_edit', user_id=user.id)

    # Pass current groups, roles, and permissions to the template
    groups = Group.objects.all()
    return render(request, 'users/user_edit.html', {
        'user': user,
        'groups': groups,
        'permissions': permissions,
        'user_permissions': user_permissions,
    })

def toggle_user_status(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        status = request.POST.get('status')
        
        try:
            user = User.objects.get(pk=user_id)
            user.is_active = (status == 'active')
            user.save()
            messages.success(request, f'User is status changed to {status}')
            return redirect('user_detail', user_id)
        except User.DoesNotExist:
            messages.error(request, 'Cannot able to change user status')
            return redirect('user_detail', user_id)
    
    return JsonResponse({'success': False, 'error': 'Invalid request.'}, status=400)

@login_required
@role_required(['super_admin','admin'])
def delete_user(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        print(user_id)
        user = get_object_or_404(User, id=user_id)

        # Check if the user is trying to delete themselves (optional check)
        if user == request.user:
            messages.success(request, 'You cannot Delete your Account by your self')
            return redirect('user_list')

        # Check if the user is assigned to a task
        if Task.objects.filter(assigned_to=user).exists():
            messages.error(request, 'This user cannot be deleted because they are assigned to a task.')
            return redirect('user_list')

        # Check if the user is the leader of any team
        if Team.objects.filter(leader=user).exists():
            messages.error(request, 'This user cannot be deleted because they are a team leader.')
            return redirect('user_list')

        # Log user deletion activity
        ActivityLog.objects.create(
            user=request.user,
            activity_type='delete_user',
            activity_description=f'Deleted user {user.username}',
            object_name=user.username,
            object_id=user.id,
            timestamp=now()
        )

        # Delete the user
        user.delete()

        # Redirect to the user list view after deletion
        messages.success(request, f'User {user.username} has been deleted successfully.')
        return redirect('user_list')

    return HttpResponseForbidden("Invalid request.")

@login_required
@role_required(['super_admin','admin'])
def user_detail(request, user_id):
    user = get_object_or_404(User, id=user_id)

    # Fetch associated data
    teams = Team.objects.filter(members=user)
    tasks = Task.objects.filter(assigned_to=user).select_related('project')
    task_projects = Project.objects.filter(tasks__assigned_to=user).distinct()
    team_projects = Project.objects.filter(assigned_team__members=user).distinct()
    projects = (task_projects | team_projects).distinct()

    # Fetch activity logs for this user
    activities = ActivityLog.objects.filter(user=user).order_by('-timestamp')

    context = {
        'user': user,
        'teams': teams,
        'tasks': tasks,
        'projects': projects,
        'activities': activities,
    }
    return render(request, 'users/user_detail.html', context)


class CustomPasswordResetCompleteView(PasswordResetCompleteView):
    def get(self, *args, **kwargs):
        """Override the GET method to skip default rendering"""
        # Add a success message
        messages.success(self.request, "Your password has been successfully reset. Please log in with your new password.")

        # Redirect to the login page
        return redirect(reverse_lazy('login'))  # Replace 'login' with the correct name of your login URL if different 


def password_reset(request):
    if request.method == "POST":
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            # Now we can safely access cleaned_data
            email = form.cleaned_data.get('email')

            # Check if the email exists in the system
            if User.objects.filter(email=email).exists():
                # Send the reset email
                form.save(request=request)
                messages.success(request, "Password reset link has been sent to your email.")
                return redirect('login')  # Redirect without 'users' namespace if needed
            else:
                # Email doesn't exist, show an error message
                messages.error(request, "The email address provided does not exist.")
        else:
            # Form is not valid
            messages.error(request, "Invalid form submission.")
    else:
        form = PasswordResetForm()

    return render(request, 'users/password_reset.html', {'form': form})

logger = logging.getLogger(__name__)

# Project Views
@login_required
def project_list(request):
    # Retrieve search query and filter from request
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', 'all')

    # Get the current date for calculating remaining days
    today = timezone.now().date()

    # Base queryset
    projects = Project.objects.all()

    # Apply search filter if a query is provided
    if search_query:
        projects = projects.filter(name__icontains=search_query)

    # Apply status filter
    if status_filter == 'completed':
        projects = projects.filter(status='completed')
    elif status_filter == 'in-progress':
        projects = projects.filter(status='ongoing', progress__lt=100)
    elif status_filter == 'pending':
        projects = projects.filter(progress=0)  # Assuming pending means no progress
    elif status_filter == 'deletion-request':
        projects = projects.filter(deletion_requested_by__isnull=False)

    # Annotate additional details for projects
    for project in projects:
        # Ensure start_date and end_date are valid dates
        if project.start_date and project.end_date:
            # Handle valid date calculations for project length and remaining days
            project.length = project.project_length()
            project.remaining_days = project.days_left()

            # Calculate progress
            project.progress = project.calculate_progress()

        else:
            # If either of the dates are invalid, set placeholders
            project.length = 0
            project.remaining_days = 0
            project.progress = 0

        # Get tasks count data if applicable
        tasks = project.tasks.all()
        project.total_tasks = tasks.count()
        project.pending_tasks = tasks.filter(status="pending").count()
        project.in_progress_tasks = tasks.filter(status="in_progress").count()
        project.completed_tasks = tasks.filter(status="completed").count()
        
    context = {
        'projects': projects,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    return render(request, 'projects/project_list.html', context)

def project_details(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    
    # Get related information
    team = project.assigned_team
    tasks = project.tasks.all()
    completed_tasks = tasks.filter(status='completed')
    
    # Get today's date
    today = date.today()
    
    # Overdue tasks: tasks where the end date is before today and their status is not "completed"
    overdue_tasks = tasks.filter(end_date__lt=today, status__in=['pending', 'in_progress'])
    
    # Count task related information
    total_tasks = tasks.count()
    completed_task_count = completed_tasks.count()
    overdue_task_count = overdue_tasks.count()

    # Get activity log for the project
    activity_logs = ActivityLog.objects.filter(object_name=project.name).order_by('-timestamp')

    context = {
        'project': project,
        'team': team,
        'total_tasks': total_tasks,
        'completed_task_count': completed_task_count,
        'overdue_task_count': overdue_task_count,
        'activity_logs': activity_logs,
        'tasks': tasks,
    }
    return render(request, 'projects/project_detail.html', context)


EXTENSION_ICON_MAP = {
    'pdf': 'bi-file-earmark-pdf',
    'jpg': 'bi-file-earmark-image', 
    'jpeg': 'bi-file-earmark-image',  # Add .jpeg to the map
    'png': 'bi-file-earmark-image',
    'doc': 'bi-file-earmark-word',
    'docx': 'bi-file-earmark-word',
    'xlsx': 'bi-file-earmark-spreadsheet',  
    'pptx': 'bi-file-earmark-slides',  
    'txt': 'bi-file-earmark-text',
    'zip': 'bi-file-earmark-zip', 
    'rar': 'bi-file-earmark-zip', 
    'tar': 'bi-file-earmark-zip',
    'gz': 'bi-file-earmark-zip',      
    'json': 'bi-file-earmark-text',  # Add .json extension
    # Add any additional file extensions here as needed
}

def file_manager(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    base_dir = os.path.join(settings.MEDIA_ROOT, 'project_files', f'Project_{project.id}')
    os.makedirs(base_dir, exist_ok=True)

    # Safeguard current_folder to avoid passing NoneType to os.path.join
    current_folder = request.GET.get('folder', '')  # Get folder or default to empty string
    if not current_folder:
        current_folder = ''  # Fallback to root if no folder is specified

    current_dir = os.path.abspath(os.path.join(base_dir, current_folder))
     # Calculate parent folder
    parent_folder_path = os.path.dirname(current_folder)  # This gives the parent folder path
    parent_folder_name = os.path.basename(parent_folder_path) if parent_folder_path else ''  # This gives the parent folder name
    # Breadcrumb Navigation
    breadcrumb = []
    path_accumulator = ""

    # Normalize the path to always use forward slashes
    normalized_folder_path = current_folder.replace("\\", "/")

    for part in normalized_folder_path.split('/'):
        if part:
            path_accumulator = os.path.join(path_accumulator, part).replace("\\", "/")
            breadcrumb.append({"name": part, "path": f"{request.path}?folder={path_accumulator}"})

    # Handle File Upload
    if request.method == "POST" and request.FILES:
        if not request.user.has_perm("gantt_app.user_can_upload_file"):
            print(f"User {request.user.username} does not have upload permission.")
            messages.error(request, "You do not have permission to upload files.")
            return redirect(f'{request.path}?folder={current_folder}')
        uploaded_file = request.FILES.get('file')
        if uploaded_file:
            file_path = os.path.join(current_dir, uploaded_file.name)  # Ensure that file is uploaded into the current folder
            with open(file_path, 'wb+') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)
            messages.success(request, f"File '{uploaded_file.name}' uploaded successfully!")
            return redirect(f'{request.path}?folder={current_folder}')  # Redirect to correct folder

    # Handle Folder Creation
    if request.method == "POST" and "folder_name" in request.POST:
        if not request.user.has_perm("gantt_app.user_can_create_folder"):
            messages.error(request, "You do not have permission to create folders.")
            return redirect(f'{request.path}?folder={current_folder}')
        folder_name = request.POST.get('folder_name', '').strip()
        if folder_name:
            folder_path = os.path.join(current_dir, folder_name)
            os.makedirs(folder_path, exist_ok=True)
            messages.success(request, f"Folder '{folder_name}' created successfully!")
            return redirect(f'{request.path}?folder={current_folder}')

    # Handle Renaming
    if request.method == "POST" and "rename" in request.POST:
        if not request.user.has_perm("gantt_app.user_can_rename_file"):
            messages.error(request, "You do not have permission to rename files.")
            return redirect(f'{request.path}?folder={current_folder}')
        old_name = request.POST.get('old_name')
        new_name = request.POST.get('new_name')
        if old_name and new_name:
            old_path = os.path.join(current_dir, old_name)
            new_path = os.path.join(current_dir, new_name)
            try:
                os.rename(old_path, new_path)
                messages.success(request, f"Renamed '{old_name}' to '{new_name}'!")
            except FileNotFoundError:
                messages.error(request, f"File or folder '{old_name}' not found.")
        else:
            messages.error(request, "Invalid file/folder name(s).")
        return redirect(f'{request.path}?folder={current_folder}')

   # Handle Moving
    if request.method == "POST" and "move" in request.POST:
        file_or_folder = request.POST.get('file_or_folder')
        target_folder = request.POST.get('target_folder')
        
        if file_or_folder and target_folder:
            # If the target folder is the parent folder, adjust the target folder path
            if target_folder == parent_folder_path:
                # Adjust the target folder path to the parent
                target_folder = os.path.dirname(current_folder)  # Remove the last folder from current folder path
            
            old_path = os.path.join(current_dir, file_or_folder)
            new_path = os.path.join(base_dir, target_folder, file_or_folder)
            
            try:
                # Ensure the target folder exists before moving
                if not os.path.exists(os.path.join(base_dir, target_folder)):
                    os.makedirs(os.path.join(base_dir, target_folder))
                
                shutil.move(old_path, new_path)
                messages.success(request, f"Moved '{file_or_folder}' to '{target_folder}'.")
            except FileNotFoundError:
                messages.error(request, f"File or folder '{file_or_folder}' not found.")
            except Exception as e:
                messages.error(request, f"Error while moving: {str(e)}")
        else:
            messages.error(request, "Please provide both file/folder and target folder.")
        return redirect(f'{request.path}?folder={current_folder}')


    # Handle Deleting
    if request.method == "POST" and "delete" in request.POST:
        if not request.user.has_perm("gantt_app.user_can_delete_file"):
            messages.error(request, "You do not have permission to delete files.")
            return redirect(f'{request.path}?folder={current_folder}')
        file_or_folder = request.POST.get('file_or_folder')
        path_to_delete = os.path.join(current_dir, file_or_folder)
        try:
            if os.path.isdir(path_to_delete):
                shutil.rmtree(path_to_delete)
            else:
                os.remove(path_to_delete)
            messages.success(request, f"Deleted '{file_or_folder}'.")
        except FileNotFoundError:
            messages.error(request, f"File or folder '{file_or_folder}' not found.")
        return redirect(f'{request.path}?folder={current_folder}')

    # Fetch directory contents
    try:
        entries = os.listdir(current_dir)
        folders = [{"name": e, "path": os.path.join(current_folder, e)} for e in entries if os.path.isdir(os.path.join(current_dir, e))]
        files = [
            {
                "name": e,
                "path": os.path.join(current_folder, e),
                "extension": e.split('.')[-1].lower() if '.' in e else '', 
                "icon": EXTENSION_ICON_MAP.get(e.split('.')[-1].lower() if '.' in e else '', 'bi bi-file-earmark'),
            }
            for e in entries if os.path.isfile(os.path.join(current_dir, e))
        ]
    except FileNotFoundError:
        raise Http404("Folder not found.")

    return render(request, 'projects/file_manager.html', {
        'project': project,
        'breadcrumb': breadcrumb,
        'folders': folders,
        'files': files,
        'parent_folder_name':parent_folder_name,
        'parent_folder_path':parent_folder_path
    })



@login_required
@role_required(['super_admin', 'admin'])
def project_create(request):
    title = "Create"
    if request.method == 'POST':
        name = request.POST.get('name')
        start_date = parse_date(request.POST.get('start_date'))
        end_date = parse_date(request.POST.get('end_date'))
        can_be_extended = request.POST.get('can_be_extended') == 'on'
        team_id = request.POST.get('team')  # Get the selected team ID
        description = request.POST.get('description')
        
        # Validate the data
        if not name or not start_date or not end_date or not team_id:
            teams = Team.objects.all()  # Fetch teams again in case of error
            return render(request, 'projects/project_form.html', {'teams': teams, 'title': title})

        if start_date > end_date:
            teams = Team.objects.all()  # Fetch teams again in case of error
            return render(request, 'projects/project_form.html', {'teams': teams, 'title': title})

        # Check if the project name already exists
        if Project.objects.filter(name=name).exists():
            messages.error(request, f"'{name}' has already been created.")
            teams = Team.objects.all()  # Fetch teams again in case of error
            return render(request, 'projects/project_form.html', {'teams': teams, 'title': title})

        try:
            team = Team.objects.get(id=team_id)
        except Team.DoesNotExist:
            messages.error(request, "Team does not exist.")
            return render(request, 'projects/project_form.html', {'teams': teams, 'title':title})

        # Create the project (without assigned_team) first
        project = Project.objects.create(
            name=name,
            start_date=start_date,
            end_date=end_date,
            description=description,
           
        )

        # Now, assign the team to the project (foreign key needs the project to be saved first)
        project.assigned_team = team
        project.save()  # Save the project again to reflect the assigned team

        # Log project creation activity
        ActivityLog.objects.create(
            user=request.user,
            activity_type='create_project',
            activity_description=f'Created project {project.name} with team {project.assigned_team.name}',
            object_name=project.name,
            object_id=project.id,
            timestamp=now()
        )

        return redirect('project_list')

    # If GET request or form validation fails
    teams = Team.objects.all()  # Fetch all available teams
    return render(request, 'projects/project_form.html', {'teams': teams, 'title': title})

@login_required
@role_required(['super_admin','admin'])
def project_edit(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    title = "Edit"
    # Fetch activity logs for this specific project
    activity_logs = ActivityLog.objects.filter(object_id=project.id, activity_type='edit_project')

    # For each activity log, extract old details
    for log in activity_logs:
        match = re.search(r'Old details: Name: (.*?), Start Date: (.*?), End Date: (.*?), Team: (.*)', log.activity_description)
        if match:
            log.old_name = match.group(1)
            log.old_start_date = match.group(2)
            log.old_end_date = match.group(3)
            log.old_team = match.group(4)

    if request.method == 'POST':
        name = request.POST.get('name')
        start_date = parse_date(request.POST.get('start_date'))
        end_date = parse_date(request.POST.get('end_date'))
        can_be_extended = request.POST.get('can_be_extended') == 'on'
        team_id = request.POST.get('team')
        description = request.POST.get('description')

        # Validate that required fields are present
        if not name or not start_date or not end_date or not team_id:
            messages.error(request, 'Name, start date, end date, and team are required!')
            return render(request, 'projects/project_form.html', {
                'project': project,
                'teams': Team.objects.all(),
                'error': 'Name, start date, end date, and team are required!'
            })

        if start_date > end_date:
            messages.error(request, 'Start date cannot be later than end date!')
            return render(request, 'projects/project_form.html', {
                'project': project,
                'teams': Team.objects.all(),
                'error': 'Start date cannot be later than end date!'
            })

        # Fetch the selected team
        team = Team.objects.get(id=team_id)

        # Save only the changed fields
        old_project_details = ""
        changes_made = False
        changed_fields = []

        # If project name is changed
        if project.name != name:
            old_project_details += f"Name: {project.name} -> {name}, "
            project.name = name
            changed_fields.append("Name")
            changes_made = True

        # If start date is changed
        if project.start_date != start_date:
            old_project_details += f"Start Date: {project.start_date} -> {start_date}, "
            project.start_date = start_date
            changed_fields.append("Start Date")
            changes_made = True

        # If end date is changed
        if project.end_date != end_date:
            old_project_details += f"End Date: {project.end_date} -> {end_date}, "
            project.end_date = end_date
            changed_fields.append("End Date")
            changes_made = True

        # If assigned team is changed
        if project.assigned_team != team:
            old_project_details += f"Team: {project.assigned_team.name if project.assigned_team else 'None'} -> {team.name}, "
            project.assigned_team = team
            changed_fields.append("Team")
            changes_made = True

        # If description is changed
        if project.description != description:
            old_project_details += f"Description: {project.description} -> {description}, "
            project.description = description
            changed_fields.append("Description")
            changes_made = True

        # If no changes were made
        if not changes_made:
            messages.info(request, 'No updates were made to the project.')
            return redirect('project_list')

        # Save the project after changes
        project.can_be_extended = can_be_extended
        project.save()

        # Log the project update activity
        changes_str = ', '.join(changed_fields)
        ActivityLog.objects.create(
            user=request.user,
            activity_type='edit_project',
            activity_description=f'Edited project {project.name}. Changed: {changes_str}. Old details: {old_project_details}',
            object_name=project.name,
            object_id=project.id,
            timestamp=timezone.now()
        )

        # Success message after update
        messages.success(request, f'Project updated successfully! Changed: {", ".join(changed_fields)}')

        return redirect('project_list')

    teams = Team.objects.all()  # Fetch all available teams
    return render(request, 'projects/project_form.html', {
        'project': project,
        'teams': teams,
        'title': title,
        'activity_logs': activity_logs,
    })
@login_required
@role_required(['super_admin','admin'])
def project_delete(request, project_id):
    project = get_object_or_404(Project, id=project_id)

    if request.method == 'POST':
        if not project.deletion_requested_by:
            # Step 1: First admin requests deletion
            project.deletion_requested_by = request.user
            project.save()
            messages.info(request, "Deletion has been requested. A second admin needs to confirm.")
            return redirect('project_list')

        elif project.deletion_requested_by != request.user:
            # Step 2: Second admin confirms deletion
            # Log project deletion activity
            ActivityLog.objects.create(
                user=request.user,
                activity_type='delete_project',
                activity_description=f'Confirmed deletion of project {project.name}',
                object_name=project.name,
                object_id=project.id,
                timestamp=now()
            )

            project.delete()
            messages.success(request, "Project has been deleted after confirmation by a second admin.")
            return redirect('project_list')

        else:
            # Prevent the same admin from confirming their own request
            messages.error(request, "You cannot confirm the deletion you initiated.")
            return redirect('project_list')

    return render(request, 'projects/project_confirm_delete.html', {'project': project})

@role_required(['super_admin','admin'])
def cancel_project_deletion(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if request.method == "POST":
        # Ensure only the user who requested deletion can cancel it
        if project.deletion_requested_by == request.user:
            # Clear the deletion request
            project.deletion_requested_by = None
            project.save()

            # Log cancelation activity
            ActivityLog.objects.create(
                user=request.user,
                activity_type="cancel_deletion_request",
                activity_description=f"Canceled deletion request for project {project.name}",
                object_name=project.name,
                object_id=project.id,
                timestamp=now()
            )
            
            messages.success(request, f"Deletion request for '{project.name}' has been canceled.")
        else:
            messages.error(request, "You cannot cancel this deletion request.")
            
    return redirect("project_list")
# Task Views
@login_required
def task_list(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    tasks = project.tasks.all()  # Fetch all tasks related to the project
    
    today = now().date()  # Get today's date
    task_list_data = []

    for task in tasks:
        status_display = dict(Task.STATUS_CHOICES).get(task.status, task.status)
        # Calculate remaining days
        days_left = (task.end_date - today).days if task.end_date and task.end_date > today else None

        # Calculate delayed days (if the task end_date is before today)
        delayed_days = (today - task.end_date).days if task.end_date and task.end_date < today else None

        # Prepare task details for the template
        task_list_data.append({
            "id": task.id,
            "name": task.name,
            "description": task.description or "No description provided",
            "start_date": task.start_date,
            "end_date": task.end_date,
            'status': status_display,
            "days_left": f"{days_left} days " if days_left else "No days",
            "delayed_days": f"{delayed_days} days delayed" if delayed_days else None,
            "assigned_to": [user.username for user in task.assigned_to.all()],  # List of usernames
        })

    return render(request, "tasks/task_list.html", {
        "project": project,
        "tasks": task_list_data,  # Send properly structured task data
    })

@login_required
@role_required(['super_admin', 'admin', 'staff'])
@project_leader_required
def task_create(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    team = project.assigned_team  # Get the team assigned to the project

    if team:
        team_members = team.members.all()
    else:
        team_members = User.objects.none()  # No team assigned

    if request.method == 'POST':
        name = request.POST.get('name')
        start_date = parse_date(request.POST.get('start_date'))
        end_date = parse_date(request.POST.get('end_date'))
        can_be_extended = request.POST.get('can_be_extended') == 'on'
        assigned_to_ids = request.POST.getlist('assigned_to')
        status = request.POST.get('status')
        description = request.POST.get('description')
        
        # Validation checks
        if not name or not start_date or not end_date:
            return render(request, 'tasks/task_form.html', {
                'error': 'Name, start date, and end date are required!',
                'project': project,
                'users': team_members,
            })

        if start_date > end_date:
            return render(request, 'tasks/task_form.html', {
                'error': 'Start date cannot be later than end date!',
                'project': project,
                'users': team_members,
            })

        # **Additional Validation for Task Dates within Project Duration**
        if start_date < project.start_date or end_date > project.end_date:
            messages.error(request, 'Task should be in a project time scope')
            return render(request, 'tasks/task_form.html', {
                'project': project,
                'users': team_members,
            })

        # Check if a task with the same name already exists for the current project
        if Task.objects.filter(name=name, project=project).exists():
            messages.error(request, f"'{name}' task already exists for this project.")
            return render(request, 'tasks/task_form.html', {
                'project': project,
                'users': team_members,
            })

        # **New Validation for Active Users**
        inactive_users = User.objects.filter(id__in=assigned_to_ids, is_active=False)
        if inactive_users.exists():
            messages.error(request, f"You cannot assign tasks to inactive users: {', '.join(user.username for user in inactive_users)}.")
            return render(request, 'tasks/task_form.html', {
                'project': project,
                'users': team_members,
            })
        
        # Create the task after validation
        task = Task.objects.create(
            name=name,
            start_date=start_date,
            end_date=end_date,
            can_be_extended=can_be_extended,
            status=status,
            description=description,
            project=project,
        )

        # Assign users to the task
        assigned_to = User.objects.filter(id__in=assigned_to_ids)
        task.assigned_to.set(assigned_to)

        # Log the task creation activity
        ActivityLog.objects.create(
            user=request.user,
            activity_type='task_created',
            activity_description=f'Created task "{task.name}" for project "{project.name}".',
            object_name=task.name,
            object_id=task.id,
            timestamp=timezone.now(),
        )

        # Redirect to the task list for this project
        messages.success(request, f'Task "{task.name}" created successfully.')
        project.update_status_based_on_tasks()  # Update the status based on tasks
        project.save()
        return redirect('task_list', project_id=project.id)

    return render(request, 'tasks/task_form.html', {
        'project': project,
        'users': team_members,
    })


@role_required(['super_admin', 'admin', 'staff'])
@project_leader_required
def task_edit(request, project_id, task_id):
    project = get_object_or_404(Project, id=project_id)
    task = get_object_or_404(Task, id=task_id, project=project)
    team = project.assigned_team
    team_members = team.members.all() if team else User.objects.none()

    if request.method == 'POST':
        # Capture original details for comparison
        original_details = {
            "name": task.name,
            "start_date": task.start_date,
            "end_date": task.end_date,
            "can_be_extended": task.can_be_extended,
            "status": task.status,
            "description": task.description,
            "assigned_to_ids": list(task.assigned_to.values_list("id", flat=True))
        }

        # Fetch updated details from the form
        updated_details = {
            "name": request.POST.get('name'),
            "start_date": parse_date(request.POST.get('start_date')),
            "end_date": parse_date(request.POST.get('end_date')),
            "can_be_extended": request.POST.get('can_be_extended') == 'on',
            "status": request.POST.get('status'),
            "description": request.POST.get('description'),
            "assigned_to_ids": list(map(int, request.POST.getlist('assigned_to'))),
        }

        # Detect changes
        changes = {}
        for field, original_value in original_details.items():
            updated_value = updated_details[field]
            if original_value != updated_value:
                changes[field] = {"old": original_value, "new": updated_value}

        # If no changes detected, skip saving
        if not changes:
            messages.info(request, "No changes were detected.")
            return redirect('task_list', project_id=project.id)

        # Update only the changed fields
        if "name" in changes:
            task.name = updated_details["name"]
        if "start_date" in changes:
            task.start_date = updated_details["start_date"]
        if "end_date" in changes:
            task.end_date = updated_details["end_date"]
        if "can_be_extended" in changes:
            task.can_be_extended = updated_details["can_be_extended"]
        if "status" in changes:
            task.status = updated_details["status"]
        if "description" in changes:
            task.description = updated_details["description"]

        # Handle assigned users only if changed
        if "assigned_to_ids" in changes:
            task.assigned_to.clear()
            if updated_details["assigned_to_ids"]:
                assigned_users = User.objects.filter(id__in=updated_details["assigned_to_ids"])
                task.assigned_to.add(*assigned_users)

        # Validate required fields
        if not task.name or not task.start_date or not task.end_date:
            return render(request, 'tasks/task_form.html', {
                'error': 'Name, start date, and end date are required!',
                'task': task,
                'project': project,
                'users': team_members,
            })

        # Validate date range
        if task.start_date > task.end_date:
            return render(request, 'tasks/task_form.html', {
                'error': 'Start date cannot be later than end date!',
                'task': task,
                'project': project,
                'users': team_members,
            })

        # Save the task with updated fields
        task.save()
        project.update_status_based_on_tasks()  # Update the status based on tasks
        project.save()
        # Log changes in ActivityLog
        change_description = "; ".join([f"{key}: {values['old']} -> {values['new']}" for key, values in changes.items()])
        ActivityLog.objects.create(
            user=request.user,
            activity_type='task_edited',
            activity_description=f'Edited task "{task.name}" for project "{project.name}". Changes: {change_description}.',
            object_name=task.name,
            object_id=task.id,
            timestamp=timezone.now(),
        )

        messages.success(request, "Task updated successfully.")
        return redirect('task_list', project_id=project.id)

    return render(request, 'tasks/task_form.html', {'task': task, 'project': project, 'users': team_members})




@login_required
@role_required(['super_admin', 'admin', 'staff'])
@project_leader_required
def task_delete(request, project_id, task_id):
    project = get_object_or_404(Project, id=project_id)
    task = get_object_or_404(Task, id=task_id, project=project)

    if request.method == 'POST':
        # Log the task deletion activity
        ActivityLog.objects.create(
            user=request.user,
            activity_type='task_deleted',
            activity_description=f'Deleted task "{task.name}" from project "{project.name}".',
            object_name=task.name,
            object_id=task.id,
            timestamp=timezone.now(),
        )

        messages.success(request, f'Task "{task.name}" Deleted successfully.')
        task.delete()
        return redirect('task_list', project_id=project.id)

   

def task_status_update(request, task_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            new_status = data.get("status")

            task = Task.objects.get(id=task_id)
            old_status = task.status  # Record the old status before updating
            task.status = new_status
            task.save(update_fields=["status"])
            project = get_object_or_404(Project, id=task.project.id)
            project.update_status_based_on_tasks()  # Update the status based on tasks
            project.save()
            # Log task status update
            ActivityLog.objects.create(
                user=request.user,
                activity_type='task_status_updated',
                activity_description=f'Updated task "{task.name}" status from "{old_status}" to "{new_status}".',
                object_name=task.name,
                object_id=task.id,
                timestamp=timezone.now(),
            )

            return JsonResponse({"success": True, "new_status": task.get_status_display()})

        except Task.DoesNotExist:
            return JsonResponse({"success": False, "error": "Task not found"})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})

@team_member_required
def task_updates(request, project_id, task_id):
    task = get_object_or_404(Task, id=task_id, project__id=project_id)
    updates = TaskUpdate.objects.filter(task=task).order_by("-date")
    
    # Determine the origin path based on the referrer or explicit request parameter
    if "projects" in request.path:
        origin_path = "projects"
    elif "dashboard" in request.META.get("HTTP_REFERER", ""):  # Use HTTP_REFERER to infer navigation
        origin_path = "dashboard"
    else:
        origin_path = None  # Default to no origin if unclear

    if request.method == "POST":
        progress_detail = request.POST.get("progress_detail")
        issues = request.POST.get("issues")
        status = request.POST.get("status")
        current_date = timezone.localtime().date()  # Ensure comparison against current local date

        # Check if an update already exists for today
        if TaskUpdate.objects.filter(task=task, date=current_date).exists():
            messages.error(request, "An update for today already exists.")
        else:
            # Create a new TaskUpdate instance
            TaskUpdate.objects.create(
                task=task,
                updated_by=request.user,
                progress_detail=progress_detail,
                issues=issues,
                date=timezone.now()
            )

            # Update Task status if it has changed
            if status and status != task.status:
                task.status = status
                task.save(update_fields=["status"])
                project = get_object_or_404(Project, id=project_id)
                project.update_status_based_on_tasks()  # Update the status based on tasks
                project.save()
                messages.success(request, "Status updated successfully!")

            messages.success(request, "Daily update and status saved successfully!")
        
        return redirect("task_updates", project_id=project_id, task_id=task_id)

    context = {
        "task": task,
        "updates": updates,
        "project_id": project_id,
        "origin_path": origin_path,  # Include origin path in context
    }
    return render(request, "tasks/task_updates.html", context)



@login_required
@role_required(['super_admin','admin'])
def team_create(request):
    search_query = request.GET.get('search', '')  # Get the search query if provided

    # Fetch all users excluding inactive and viewer role users
    users = User.objects.filter(is_active=True).exclude(role__in=['viewer', 'admin', 'super_admin'])
    searched_user = users

    # Filter users based on the search query if applicable
    if search_query:
        searched_user = users.filter(username__icontains=search_query)

    # Prepare team data for each user
    users_with_team_data = []
    for user in searched_user:
        teams_led = Team.objects.filter(leader=user)
        team_count = Team.objects.filter(members=user).count()
        users_with_team_data.append({
            "username": user.username,
            "teams_led": teams_led,
            "team_count": team_count,
        })

    # Pagination
    paginator = Paginator(users_with_team_data, 4)  # Show 4 users per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        name = request.POST.get('name')
        team_leader_id = request.POST.get('team_leader')
        team_members_ids = request.POST.getlist('team_members')

        # Validate required fields
        if not name or not team_leader_id:
            messages.error(request, 'Team name and leader are required.')
            return redirect('team_create')
        
        # Check if the team name already exists
        if Team.objects.filter(name=name).exists():
            messages.error(request, f"'{name}' is already been created.")
            return redirect('team_create')

        try:
            # Get the leader user and validate they are not inactive or a viewer
            leader = get_object_or_404(User.objects.filter(is_active=True).exclude(role='viewer'), id=team_leader_id)

            # Create the team
            team = Team.objects.create(name=name, leader=leader)

            # Add members to the team, excluding inactive or viewer role members
            members = User.objects.filter(id__in=team_members_ids, is_active=True).exclude(role='viewer')
            team.members.set(members)
            team.members.add(leader)  # Ensure leader is also a member

         

            messages.success(request, f'Team {team.name} created successfully.')
            return redirect('team_list')
        except Exception as e:
            messages.error(request, f"Error creating team: {str(e)}")
            return redirect('team_create')

    # GET request
    return render(
        request,
        'teams/team_create.html',
        {
            'users': users,
            'searched_user': searched_user,
            'page_obj': page_obj,
            'search_query': search_query,
        },
    )


@login_required
@role_required(['super_admin','admin'])
def team_list(request):
    # Get the search query from the request
    search_query = request.GET.get('search', '')

    # Filter teams by name if there's a search query
    if search_query:
        teams = Team.objects.filter(name__icontains=search_query)
    else:
        teams = Team.objects.all()

    # Paginate the results
    paginator = Paginator(teams, 5)  # Show 5 teams per page

    # Get the page number from the GET parameter, or default to 1 if not present or invalid
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.get_page(page_number)
    except ValueError:
        raise Http404("Page not found.")  # Handle the case when the page number is invalid

    return render(request, 'teams/team_list.html', {
        'page_obj': page_obj,
        'search_query': search_query,
    })
@login_required
@role_required(['super_admin','admin','staff'])
def add_team_member(request, team_id):
    team = get_object_or_404(Team, id=team_id)

    if request.method == 'POST':
        member_id = request.POST.get('member')
        member = User.objects.get(id=member_id)

        if member in team.members.all():
            messages.error(request, "This user is already a team member.")
        else:
            # Add the member to the team
            team.members.add(member)
            messages.success(request, f"{member.username} has been added to the team.")

            # Create Activity Log for the new member addition
            ActivityLog.objects.create(
                user=request.user,  # The user who is adding the team member
                activity_type='team_member_added',  # Activity type for adding member
                activity_description=f'{member.username} has been added to the team "{team.name}".',  # Description
                object_name=team.name,  # Name of the team
                object_id=team.id,  # ID of the team
                timestamp=timezone.now(),  # Timestamp of the log
            )

        return redirect('team_detail', team_id=team.id)

    # Exclude users already in the team
    available_users = User.objects.exclude(id__in=team.members.all())
    return render(request, 'teams/add_member.html', {'team': team, 'available_users': available_users})

@login_required
@role_required(['super_admin','admin','staff'])
def edit_team(request, team_id):
    team = get_object_or_404(Team, id=team_id)

    if team.leader and team.leader not in team.members.all():
        team.members.add(team.leader)
        team.save()

    all_users = User.objects.filter(is_active=True).exclude(role__in=['viewer', 'admin', 'super_admin'])
    

    if request.method == "POST":
        # Update Team Details
        name = request.POST.get("name")
        description = request.POST.get("description")
        if name:
            old_name = team.name
            team.name = name
            ActivityLog.objects.create(
                user=request.user,
                activity_type='team_updated',
                activity_description=f'Team name changed from "{old_name}" to "{name}".',
                object_name=team.name,
                object_id=team.id,
                timestamp=timezone.now(),
            )
        if description:
            team.description = description

        # Replace Team Leader
        new_leader_id = request.POST.get("new_leader")
        if new_leader_id:
            new_leader = get_object_or_404(User, id=new_leader_id)
            if new_leader in team.members.all():
                old_leader = team.leader
                team.leader = new_leader
                ActivityLog.objects.create(
                    user=request.user,
                    activity_type='team_leader_changed',
                    activity_description=f'Leader changed from {old_leader.username} to {new_leader.username}.',
                    object_name=team.name,
                    object_id=team.id,
                    timestamp=timezone.now(),
                )
            else:
                messages.error(request, f"{new_leader.username} must be a member before becoming the leader.")

       # Add Team Members
        new_member_ids = request.POST.getlist("new_members")  # Match the 'name' in the HTML form
        if new_member_ids:
            for new_member_id in new_member_ids:
                try:
                    # Retrieve the new member and validate
                    new_member = get_object_or_404(User.objects.filter(is_active=True), id=new_member_id)

                    # Add member if not already in the team
                    if new_member not in team.members.all():
                        team.members.add(new_member)
                        ActivityLog.objects.create(
                            user=request.user,
                            activity_type='team_member_added',
                            activity_description=f'{new_member.username} added to the team.',
                            object_name=team.name,
                            object_id=team.id,
                            timestamp=timezone.now(),
                        )
                        messages.success(request, f"{new_member.username} added to the team.")
                    else:
                        messages.info(request, f"{new_member.username} is already in the team.")
                except Exception as e:
                    messages.error(request, f"Error adding member {new_member_id}: {str(e)}")


        # Remove Team Member
        remove_member_id = request.POST.get("remove_member")
        if remove_member_id:
            remove_member = get_object_or_404(User, id=remove_member_id)

            # Prevent removal if the member is the team leader
            if remove_member == team.leader:
                messages.error(request, "Cannot remove the team leader as a member. Replace the leader first.")
            else:
                # Check if the user has any uncompleted tasks assigned to them
                tasks_assigned_to_remove_member = Task.objects.filter(assigned_to=remove_member, project__assigned_team=team)
                incomplete_tasks = tasks_assigned_to_remove_member.exclude(status='COMPLETED')

                if incomplete_tasks.exists():
                    messages.error(request, f"Cannot remove {remove_member.username} because they have incomplete tasks.")
                    return redirect("team_list")

                # If all tasks are completed, remove them
                team.members.remove(remove_member)
                messages.success(request, f"{remove_member.username} removed from the team.")

                ActivityLog.objects.create(
                    user=request.user,
                    activity_type='team_member_removed',
                    activity_description=f'{remove_member.username} removed from the team.',
                    object_name=team.name,
                    object_id=team.id,
                    timestamp=timezone.now(),
                )


        team.save()
        return redirect("team_list")

    context = {
        "team": team,
        "all_users": all_users,
        "team_members": team.members.all(),
    }
    return render(request, "teams/edit_team.html", context)


@login_required
@role_required(['super_admin','admin'])
def delete_team(request):
    if request.method == 'POST':
        team_id = request.POST.get('team_id')
       
        team = get_object_or_404(Team, id=team_id)
        
        if Project.objects.filter(assigned_team=team).exclude(status='completed').exists():
            messages.error(request, f"Cannot delete team '{team.name}' because it is assigned to incomplete projects.")
            return redirect('team_list')  # Redirect to the team list view

        # Create activity log for deletion
        ActivityLog.objects.create(
            user=request.user,  # The user who is deleting the team
            activity_type='team_deleted',  # The activity type
            activity_description=f'Team "{team.name}" deleted by {request.user.username}.',  # Description
            object_name=team.name,  # Name of the team
            object_id=team.id,  # ID of the team
            timestamp=now(),  # Timestamp for the log
        )

        # Delete the team
        team.delete()
        messages.success(request, f"Team '{team.name}' deleted successfully.")
        return redirect('team_list')  # Redirect to the team list view after deletion

    return HttpResponseForbidden("Invalid request.")


@login_required
@role_required(['super_admin','admin','staff'])
def remove_team_member(request, team_id, member_id):
    team = get_object_or_404(Team, id=team_id)
    member = get_object_or_404(User, id=member_id)

    if member not in team.members.all():
        messages.error(request, "This user is not a member of the team.")
    else:
        team.members.remove(member)

        # Log the removal action
        ActivityLog.objects.create(
            user=request.user,  # The user who removed the member
            activity_type='team_member_removed',  # Activity type for member removal
            activity_description=f'{member.username} was removed from the team "{team.name}".',
            object_name=team.name,  # Team name
            object_id=team.id,  # Team ID
            timestamp=timezone.now(),  # Timestamp
        )

        messages.success(request, f"{member.username} has been removed from the team.")

    return redirect('team_detail', team_id=team.id)


# View team details
@login_required
@role_required(['super_admin','admin','staff'])
def team_detail(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    return render(request, "teams/team_detail.html", {"team": team})

@login_required
@role_required(['super_admin','admin','staff'])
def replace_team_leader(request, team_id):
    # Get the team and ensure the user is allowed to replace the leader
    team = get_object_or_404(Team, id=team_id)
    current_user = request.user

    # Only allow the current team leader or admin to replace the leader
    if team.leader != current_user and not current_user.is_superuser:
        messages.error(request, "You are not authorized to change the team leader.")
        return redirect('team_list')

    if request.method == 'POST':
        new_leader_id = request.POST.get('new_leader')
        
        # Ensure the new leader is a member of the team
        new_leader = User.objects.get(id=new_leader_id)
        if new_leader not in team.members.all():
            messages.error(request, "The selected member is not a valid team member.")
            return render(request, 'replace_leader.html', {'team': team, 'members': team.members.all()})
        
        # Log activity of changing the leader
        old_leader = team.leader  # Current leader before change

        # Replace the leader
        team.leader = new_leader
        team.save()

        # Create activity log for replacing team leader
        ActivityLog.objects.create(
            user=request.user,  # User who made the change
            activity_type='team_leader_replaced',  # Activity type for leader replacement
            activity_description=f"{old_leader.username} has been replaced by {new_leader.username} as the leader of the team {team.name}.",  # Activity description
            object_name=team.name,  # Team name
            object_id=team.id,  # Team ID
            timestamp=timezone.now(),  # Timestamp of the activity
        )

        messages.success(request, f"Team leader has been updated to {new_leader.username}.")
        return redirect('team_list')

    # GET request: Display the form to replace the team leader
    return render(request, 'teams/replace_leader.html', {'team': team, 'members': team.members.all()})
